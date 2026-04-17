#!/usr/bin/env python3
"""
Генерира bg_cheapest_v3_YYYY-MM-DD.xlsx от master JSON файла.

Листи:
  1. Най-евтини по категория  — топ 5 по нормализирана цена
  2. All by category           — всички продукти по категория
  3. Обобщение                — най-евтиният от всяка подкатегория
  4. Сравнение (авто)        — Union-Find по ключови думи
  5. Сравнение (ИИ)          — Azure OpenAI GPT-4o
  6. Всички продукти          — пълен набор с обогатени полета
  7. За преглед               — некласифицирани + без единица
"""

import json
import re
import sys
import time
from collections import defaultdict
from datetime import date
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Грешка: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE         = Path(__file__).parent
MASTER_PATH  = BASE / "bulgarian_promo_prices_merged.json"
CONFIG_PATH  = BASE / "azure_config.json"
SECRETS_PATH = BASE / "azure_secrets.json"
OVERRIDES_PATH = BASE / "manual_overrides.json"
OUTPUT_PATH  = BASE / f"bg_cheapest_v4_{date.today()}.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT     = Font(name='Arial', size=10)
BOLD_FONT     = Font(name='Arial', bold=True, size=10)
TITLE_FONT    = Font(name='Arial', bold=True, size=13, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', italic=True, size=10, color='595959')
CATROW_FONT   = Font(name='Arial', bold=True, size=11, color='1F4E79')
SUBCAT_FONT   = Font(name='Arial', bold=True, size=10, color='384060')
NOTE_FONT     = Font(name='Arial', italic=True, size=10, color='808080')
GREEN_FONT    = Font(name='Arial', bold=True, size=10, color='006100')
URL_FONT      = Font(name='Arial', size=10, color='0563C1', underline='single')
MONEY_FMT     = '#,##0.00" €"'
NORM_FMT      = '#,##0.00'
PCT_FMT       = '0%'
GREEN         = PatternFill('solid', fgColor='C6EFCE')
BLUE_LIGHT    = PatternFill('solid', fgColor='D6E4F0')
YELLOW        = PatternFill('solid', fgColor='FFF2CC')
GRAY          = PatternFill('solid', fgColor='F2F2F2')
CAT_FILL      = PatternFill('solid', fgColor='D9E2F3')
SUBCAT_FILL   = PatternFill('solid', fgColor='E9EDF4')
REVIEW_FILL   = PatternFill('solid', fgColor='EBF3FB')
BORDER        = Border(bottom=Side(style='thin', color='D9D9D9'))
THICK_BORDER  = Border(bottom=Side(style='medium', color='1F4E79'))

CAT_ORDER = [
    "Млечни продукти", "Яйца", "Месо", "Колбаси и деликатеси",
    "Риба и морски дарове", "Плодове", "Зеленчуци",
    "Хляб и тестени", "Основни хранителни", "Готови храни",
    "Напитки", "Сладкарски", "Домакински",
    "Козметика и хигиена", "Домашни любимци", "Дом и градина", "Други",
]

# ── Category Rules ─────────────────────────────────────────────────────────────
# (category, subcategory, include_patterns, exclude_patterns) — first match wins
RULES = [
    # LAYER 0 — Brand Anchors & Vitrina (before all meat/cheese/vegetable rules)
    ("Домакински", "Перилни/Почистващи",
     [r'\bfrosch\b', r'\bmedix\b', r'\bmeglio\b'], [r'ароматизатор']),
    ("Домакински", "Ароматизатори",
     [r'\bglade\b'], []),
    ("Козметика и хигиена", "Шампоан/Коса",
     [r'\bgarnier\b.*(?:шамп|балсам|маска|therapy)', r'\bsyoss\b', r'\bbatiste\b',
      r'\bhair vege\b', r'\bnivea\b.*шамп', r'\bwash.*go\b'], []),
    ("Козметика и хигиена", "Крем за лице",
     [r"l'oreal", r"l'or[eé]al", r'\brevitalift\b'], [r'шамп', r'коса']),
    ("Козметика и хигиена", "Бръснене",
     [r'\bgillette\b'], []),
    ("Домашни любимци", "Котешка тоалетна",
     [r'котешка тоалетна', r'постелк.*за котк', r'пясък за котк',
      r'arena.*котк', r'whisker.*litter'], []),
    ("Домашни любимци", "Котешка храна",
     [r'k-carinura', r'храна за кот', r'лакомство за кот',
      r'стикове за кот', r'пастет за кот',
      r'\bwhiskas\b', r'\bfelix\b', r'\bsheba\b'], [r'куч']),
    ("Домашни любимци", "Кучешка храна",
     [r'\bdante\b.*куч', r'храна за куч', r'лакомство за куч',
      r'стикове за куч', r'пастет за куч',
      r'\bpedigree\b', r'\bbeneful\b'], []),
    ("Домашни любимци", "Аксесоари за животни",
     [r'нашийник', r'повод.*куч', r'играчка.*куч', r'играчка.*кот',
      r'легло.*куч', r'легло.*кот', r'транспортна клетка'], []),
    # Vitrina — MUST come before meat/cheese/vegetable rules
    ("Готови храни", "Ястия от витрина",
     [r'от топлата витрина', r'billa ready', r'от деликатесната витрина',
      r'ротисер', r'цяло печено пиле', r'топла витрина'], []),
    # LAYER 1 — Household
    ("Домакински", "Торби за смет",
     [r'торби за смет', r'торбички.*смет', r'чували за смет',
      r'торби за отпадъци', r'торбички тип потник', r'торби за тежки отпадъци',
      r'[kк]-bio.*торб', r'[nн]evos.*[tт]орб',
      r'k-classic.*торб', r'биохартиена.*торб'], []),
    ("Домакински", "Тоалетна хартия",
     [r'[tт]оалетна хартия'], []),
    ("Домакински", "Кухненска ролка",
     [r'кухненска ролка'], []),
    ("Домакински", "Салфетки",
     [r'\bсалфетки\b'], []),
    ("Домакински", "Перилни/Почистващи",
     [r'препарат', r'обезмаслител', r'гел за пране', r'прах за пране', r'омекотител',
      r'таблетки.*съдомиялн', r'гел.*съдомиялн', r'сол за съдомиялн',
      r'дезинфек', r'wc блокче', r'wc гел', r'wc аромат', r'хигиенен спрей',
      r'против петна'], []),
    ("Домакински", "Ароматизатори",
     [r'ароматизатор'], []),
    ("Домакински", "Дамски превръзки",
     [r'дамски превръзки'], []),
    # LAYER 1.5 — Cosmetics
    ("Козметика и хигиена", "Шампоан/Коса",
     [r'шампоан', r'балсам.*коса', r'маска за коса', r'серум за коса',
      r'спрей за корени', r'боя за коса', r'сух шампоан', r'\bolia\b'], []),
    ("Козметика и хигиена", "Крем за лице",
     [r'крем за лице', r'серум.*лице', r'флуид.*лице', r'age specialist',
      r'hyaluron', r'bright reveal'], []),
    ("Козметика и хигиена", "Бръснене",
     [r'бръсн', r'самобръсначка'], []),
    ("Козметика и хигиена", "Душ гел/Сапун",
     [r'душ гел', r'течен сапун', r'\bteo\b.*rich', r'\btreaclemoon\b', r'\bziaja\b'], []),
    # LAYER 2 — Food
    ("Млечни продукти", "Прясно мляко",
     [r'прясно мляко', r'адаптирано мляко', r'\bмляко\b'],
     [r'кисело', r'цедено', r'кокосово', r'ацидофилно', r'пробиотично', r'напитка', r'десерт']),
    ("Млечни продукти", "Кисело мляко",
     [r'кисело мляко', r'ацидофилно.*мляко', r'пробиотично.*мляко'], []),
    ("Млечни продукти", "Цедено мляко",
     [r'цедено.*мляко'], []),
    ("Млечни продукти", "Специални сирена",
     [r'\bрикота\b', r'грана падано', r'\bмаскарпоне\b', r'\bпекорино\b', r'\bманчего\b'], []),
    ("Млечни продукти", "Сирене",
     [r'\bсирене\b', r'\bмоцарела\b', r'\bbrimi\b', r'\bphiladelphia\b',
      r'крема сирене', r'котидж', r'\bбри\b', r'синьо сирене'], [r'панирани.*сиренца']),
    ("Млечни продукти", "Кашкавал",
     [r'\bкашкавал\b', r'\bементал\b', r'\bгауда\b.*(?:слайс|300)'], []),
    ("Млечни продукти", "Масло (краве)",
     [r'масло краве', r'markenbutter'], [r'маслиново', r'олио']),
    ("Млечни продукти", "Сметана/Извара",
     [r'заквасена сметана', r'\bизвара\b', r'крем за разбиване'], []),
    ("Млечни продукти", "Млечни напитки/десерти",
     [r'млечна напитка', r'млечен десерт', r'протеинов йогурт'], []),
    ("Яйца", "Яйца",
     [r'\bяйца\b', r'\bбиояйца\b'], [r'без яйца', r'\bуред\b']),
    # Meat — specialized subcategories BEFORE general ones
    ("Месо", "Пилешки карантии",
     [r'пилешки дробчета', r'пилешки сърца', r'пилешки стомашета',
      r'пилешки кокали', r'пилешки крилца', r'пилешки черен дроб',
      r'пилешк.*дробчет', r'пилешк.*сърц'], []),
    ("Месо", "Агнешко",
     [r'агнешк'], []),
    ("Месо", "Пилешко",
     [r'пилеш', r'пиле\b', r'\bfragedo\b', r'la provincia.*пилеш',
      r'царевично пиле', r'frangosul.*пилеш'],
     [r'кренвирш', r'супа', r'бульон', r'дробчет', r'сърц', r'стомашет',
      r'от топлата витрина', r'billa ready']),
    ("Месо", "Свинско",
     [r'свинск'], [r'салам', r'шунка', r'наденица', r'кренвирш', r'шпек', r'бекон']),
    ("Месо", "Телешко",
     [r'телеш'], [r'шкембе', r'бульон', r'супа', r'агнешк']),
    ("Месо", "Кайма",
     [r'\bкайма\b'], [r'кюфте', r'кебапче', r'кюфтета', r'кебапчета']),
    ("Месо", "Шкембе",
     [r'\bшкембе\b'], []),
    ("Колбаси и деликатеси", "Шунка",
     [r'\bшунка\b', r'\bпастърма\b'], []),
    ("Колбаси и деликатеси", "Салам/Колбас",
     [r'\bсалам\b', r'\bколбас\b', r'\bсуджук\b', r'\bлуканка\b', r'\bсушеник\b',
      r'\bнаденица\b', r'\bдебърцини\b', r'\bкренвирш', r'\bшпек\b', r'\bбекон\b',
      r'\bмезе\b', r'\bфуетек\b', r'кълцано карначе',
      r'\bпастет\b(?!.*куч)(?!.*кот)'], []),
    # Fish — canned/marinated BEFORE fresh
    ("Риба и морски дарове", "Консервирана/Маринована риба",
     [r'риба.*консерв', r'консерв.*риба', r'риба.*маринов', r'маринов.*риба',
      r'\bриба тон\b', r'\bтон\b.*риба', r'сардини', r'паламуд',
      r'шпроти', r'\btuna\b', r'\btonno\b', r'скумрия.*консерв',
      r'скумрия.*маринов', r'\bтриска\b'], []),
    ("Риба и морски дарове", "Риба",
     [r'\bриб', r'\bсьомг', r'\bскумрия\b', r'\bпъстърв', r'\bсельод', r'\bхайвер\b',
      r'\bсуши\b', r'\bмиди\b', r'\bсурими\b', r'\bocean\b', r'\bmowi\b',
      r'\bхек\b', r'филе панирано'], []),
    # Fruits — exclude sodas, drinks, vinegar, juices
    ("Плодове", "Плодове",
     [r'\bбанан', r'биобанан', r'\bябълк', r'\bягоди\b', r'\bгрозде\b', r'\bкруш',
      r'\bпортокал', r'\bлимон(?!ад)', r'\bлимет', r'\bкиви\b', r'\bборовинк',
      r'\bананас\b', r'\bкестен', r'\bсливи\b', r'\bфурми\b', r'\bавокадо\b'],
     [r'сок', r'нектар', r'смути', r'препарат', r'обезмаслител', r'чипс', r'вино',
      r'сайдер', r'напитка', r'газирана', r'безалкохолна', r'\bmirinda\b', r'\baspasia\b',
      r'\bоцет\b', r'\b100%\b', r'happy day']),
    # Vegetables — exclude vitrina, formed products, olive oil, snacks
    ("Зеленчуци", "Зеленчуци",
     [r'\bкартоф', r'\bдомат', r'\bкраставиц', r'\bчушк', r'биочушк', r'\bлук\b',
      r'\bчесън\b', r'\bтиквичк', r'\bкопфсалат', r'салата.*фризе',
      r'салата.*бон тон', r'\bспанак\b', r'\bгъби\b', r'\bцелина\b',
      r'\bрепичк', r'\bпащърнак\b', r'\bмаслин'],
     [r'чипс', r'салата снежанка', r'руска салата', r'кьопо',
      r'лютеница', r'печена капия', r'препарат',
      r'от деликатесната витрина', r'billa ready', r'от топлата витрина',
      r'картофено кюфте', r'картофени кюфтета', r'маслиново', r'\bpom bar\b']),
    ("Хляб и тестени", "Хляб",
     [r'\bхляб\b', r'\bбагет\b', r'\bземел\b', r'\bсомун\b', r'\bбиохляб\b'], []),
    ("Хляб и тестени", "Козунак",
     [r'\bкозунак\b', r'козуначен', r'\bcolomba\b'], []),
    ("Хляб и тестени", "Кроасан/Кифли",
     [r'\bкроасан\b', r'\bкифлич', r'\bпоничк', r'\bдонът\b', r'\bеклер\b'], []),
    ("Хляб и тестени", "Тесто/Кори",
     [r'\bтесто\b', r'\bкори\b', r'бутертесто', r'точени'], []),
    # Sugar — exclude drinks and "без захар" items
    ("Основни хранителни", "Захар",
     [r'\bзахар\b'],
     [r'декорация', r'захарна', r'без захар', r'безалкохолна', r'напитка',
      r'газирана', r'\b7up\b', r'\b7 up\b', r'\bpepsi\b', r'coca.cola',
      r'\bfanta\b', r'\bsprite\b', r'\bmirinda\b']),
    ("Основни хранителни", "Брашно",
     [r'\bбрашно\b', r'\bнишесте\b', r'овесени трици'], []),
    ("Основни хранителни", "Олио",
     [r'\bолио\b', r'слънчогледово'], [r'маслиново']),
    ("Основни хранителни", "Маслиново масло",
     [r'маслиново масло', r'\bbertolli\b', r"costa d.oro"], []),
    ("Основни хранителни", "Ориз/Булгур/Бобови",
     [r'\bориз\b', r'\bбулгур\b', r'\bбоб\b', r'зелен грах', r'сладка царевица'],
     [r'манастирски']),
    ("Основни хранителни", "Паста/Макарони",
     [r'\bпаста\b(?!.*зъби)', r'макаронени'], [r'без яйца']),
    ("Основни хранителни", "Оцет/Подправки",
     [r'\bкетчуп\b', r'доматено пюре', r'сода бикарбонат', r'\bоцет\b', r'балсамов',
      r'\bmaggi\b', r'зеленчуков бульон',
      r'пилешки бульон кубчета', r'бульон кубчета'], []),
    ("Основни хранителни", "Ядки и семена",
     [r'\bфъстък', r'\bфъстъц', r'\bкашу\b', r'шам фъстък', r'\bорех',
      r'\bлешник', r'\bбадем', r'слънчогледови семк', r'тиквени семк',
      r'микс.*ядк', r'\bядки\b', r'\bядков\b', r'печени ядки', r'сурови ядки'],
     [r'олио', r'масло', r'паста.*фъстък']),
    ("Основни хранителни", "Подправки и билки",
     [r'\bкуркума\b', r'\bриган\b', r'\bкимион\b', r'\bрозмарин\b', r'\bчубрица\b',
      r'\bканела\b', r'\bджинджифил\b', r'\bмащерка\b', r'\bбосилек\b',
      r'\bмагданоз\b', r'\bкопър\b', r'\bподправка\b', r'\bkotanyi\b', r'\bpicantina\b',
      r'черен пипер', r'бял пипер', r'червен пипер', r'смес подправки'],
     [r'чай', r'напитка']),
    # Prepared food — specialized subcategories FIRST
    ("Готови храни", "Супи",
     [r'крем супа', r'супа топчета', r'\bчорба\b',
      r'пилешка супа', r'телешка супа', r'зеленчукова супа',
      r'\bбульон\b(?!.*кубче)'], []),
    ("Готови храни", "Готови салати",
     [r'руска салата', r'салата снежанка', r'салата северняшка',
      r'\bтаратор\b', r'\bкьопо', r'печена капия', r'\bцацики\b'], []),
    ("Готови храни", "Готови ястия",
     [r'\bкюфтет', r'\bбургер\b', r'\bпуканки\b', r'\bкебап\b',
      r'\bпанирани\b', r'\bхумус\b', r'\bлютеница\b',
      r'\bкатък\b', r'боб по манастирски',
      r'минибанички', r'\bмекици\b', r'\bпица\b', r'\bгюбек\b',
      r'продукт за мазане', r'\bзакуска с\b',
      r'картофено кюфте', r'картофени кюфтета'], []),
    # LAYER 3 — Beverages
    ("Напитки", "Бира",
     [r'\bбира\b', r'\bсайдер\b', r'\bheineken\b', r'\bcorona\b', r'\bstaropramen\b',
      r'\bamstel\b', r'\bpaulaner\b', r'\btuborg\b', r'\bзагорка\b', r'\bкаменица\b',
      r'\bариана\b', r'\bпиринско\b', r'\bшуменско\b', r'\bбургаско\b',
      r'\bвитошко\b', r'\bболярка\b', r'\bradeberger\b', r'\bschofferhofer\b',
      r'\bsomersby\b', r'\bcorsendonk\b', r'\btemplier\b'], []),
    ("Напитки", "Вино",
     [r'\bвино\b', r'\bрозе\b(?!.*шоколад)', r'пино гриджо'], []),
    ("Напитки", "Джин/Бърбън",
     [r'\bджин\b', r'\bgin\b(?!.*ginger)', r'\bбърбън\b', r'\bбурбон\b', r'\bburbon\b',
      r'\bтекила\b', r'\bликьор\b', r'\bликер\b', r'\bконяк\b', r'\bcognac\b',
      r'\babsinthe\b', r'\bабсент\b'], []),
    ("Напитки", "Спиртни напитки",
     [r'\bводка\b', r'\bуиски\b', r'\bузо\b', r'\bгроздова\b',
      r'zacapa.*ром', r'\bром\b(?!.*аром)'], [r'\bаром', r'\bтракия']),
    ("Напитки", "Безалкохолни",
     [r'coca-cola', r'\bfanta\b', r'\bsprite\b', r'\bpepsi\b', r'\bschweppes\b',
      r'\bderby\b.*напитка', r'\bmonster\b', r'енергийна напитка',
      r'газирана напитка', r'\bлимонада\b',
      r'\bmirinda\b', r'\baspasia\b', r'\b7up\b', r'\b7 up\b'], []),
    ("Напитки", "Сокове",
     [r'\bсок\b', r'\bнектар\b', r'\bсмути\b', r'\bcappy\b', r'\bflorina\b',
      r'\bfreshko\b', r'\bhappy day\b', r'сок.*100%', r'100%.*сок',
      r'100%.*портокал', r'100%.*ябълк'], []),
    ("Напитки", "Студен чай",
     [r'студен чай', r'\bnestea\b'], []),
    ("Напитки", "Кафе",
     [r'\bкафе\b', r'\bjacobs\b', r'\bnescafe\b', r'\blavazza\b',
      r'\btchibo\b', r'\bespresso\b'], []),
    ("Напитки", "Вода",
     [r'\bdevin\b.*вод', r'горна баня', r'трапезна вода', r'минерална вода'],
     [r'филтър', r'кана']),
    # LAYER 4 — Sweets & Snacks
    ("Сладкарски", "Шоколад/Бонбони",
     [r'\bшоколад\b(?!.*фигур)', r'\bбонбон', r'\blindt\b', r'\bnutella\b',
      r'\bmoritz\b', r'\broshen\b', r'\bsnickers\b', r'\bmars\b', r'\btwix\b',
      r'\bkit kat\b', r'\blion\b.*десерт', r'\btoffifee\b'], [r'великденски заек']),
    ("Сладкарски", "Бисквити/Вафли",
     [r'\bбисквит', r'\bвафл[аи]', r'\bмаркизит', r'\b7 days\b',
      r'\bmilka\b.*бисквит', r'\bнасладки\b', r'\bтраяна\b',
      r'\bцаревец\b.*вафли', r'\bparadise\b', r'вафлен бар',
      r'\bкристал\b.*бисквит', r'бирени пръчици', r'\bbrusketi\b', r'\bmaretti\b'], []),
    ("Сладкарски", "Чипс",
     [r'\bчипс\b', r'\bdoritos\b', r'\bpom bar\b'], []),
    ("Сладкарски", "Снакс",
     [r'\bkubeti\b', r'ръжени кубчета', r'\bсолети\b', r'\bпуканки\b'], []),
    # Сладолед BEFORE Торти — profiteroles with сладолед go here, not to Торти
    ("Сладкарски", "Сладолед",
     [r'\bсладолед\b'], []),
    ("Сладкарски", "Торти/Сладкиши",
     [r'\bторта\b', r'\bчийзкейк\b', r'\bсуфле\b', r'крем брюле',
      r'\bпрофитероли\b', r'\bтирамису\b', r'\bсладкиш', r'великденски заек',
      r'\bкейк\b', r'домашни сладки', r'\bbalocco\b'],
     [r'великденска чаша', r'сладолед', r'декорация']),
    ("Сладкарски", "Кремове/Декорации за печене",
     [r'крем без варене', r'смес за печене', r'микс за брауни',
      r'декорация', r'звездички', r'\bdolce\b.*крем', r'\bdolce\b.*декор'], []),
    ("Сладкарски", "Зърнени закуски",
     [r'зърнена закуска'], []),
    # LAYER 5 — Home, Garden, Other
    ("Дом и градина", "Градина",
     [r'\bсубстрат\b', r'торфена смес', r'\bсаксия\b', r'\bгербер\b',
      r'\bсандъче\b', r'семена.*моята', r'подложка за саксия'], []),
    ("Дом и градина", "Кухненски",
     [r'\bтенджера\b', r'прибори за хранене', r'\bнож\b', r'чаши.*комплект',
      r'\bspice.*soul\b', r'\bудължител\b'], []),
    ("Други", "Играчки/Игри",
     [r'\blego\b', r'\bпъзел\b', r'дървена игра', r'топка за футбол',
      r'чадър детски', r'шоколадова фигур', r'плюшена играчка', r'великденска чаша'], []),
    ("Други", "Изкуство/Хартия",
     [r'платно за рисуване', r'блок за рисуване', r'хартия.*цвят', r'\btalentus\b'], []),
    ("Други", "Филтри за вода",
     [r'филтър за вода', r'филтри за вода', r'кана за вода'], []),
    ("Други", "Инструменти",
     [r'\bparkside\b.*станция', r'запояване'], []),
]

# ── Step 1: Unit Parser ────────────────────────────────────────────────────────
def _extract_measure(s):
    """
    Try to extract (quantity, base_unit) from a string. Returns None if no match.
    Handles: кг, г, гр (both are Bulgarian abbreviations for grams), л, мл, multipacks.
    Input must already be lowercased.
    """
    # Multipack: "NxM г/гр/мл" — e.g. "3x300 г", "2x500мл"
    m = re.search(r'(\d+)\s*[xх]\s*(\d+(?:[.,]\d+)?)\s*(гр?|мл)', s)
    if m:
        n, q = int(m.group(1)), float(m.group(2).replace(',', '.'))
        return (n * q, 'g') if m.group(3).startswith('г') else (n * q, 'ml')
    # Kilograms — must come before gram rule to avoid "кг" partially matching "г"
    m = re.search(r'([\d]+[.,]?[\d]*)\s*кг', s)
    if m:
        return float(m.group(1).replace(',', '.')) * 1000, 'g'
    # Grams: "г" or "гр" (both valid Bulgarian abbrevs — "200 г", "200 гр", "200ГР")
    # \b ensures "гр" doesn't match inside longer words; кг already handled above
    m = re.search(r'([\d]+[.,]?[\d]*)\s*гр?\b', s)
    if m:
        return float(m.group(1).replace(',', '.')), 'g'
    # Liters — (?!в) prevents matching "лв" (currency)
    m = re.search(r'([\d]+[.,]?[\d]*)\s*л(?!в)', s)
    if m:
        return float(m.group(1).replace(',', '.')) * 1000, 'ml'
    # Milliliters
    m = re.search(r'([\d]+[.,]?[\d]*)\s*мл', s)
    if m:
        return float(m.group(1).replace(',', '.')), 'ml'
    return None


def parse_unit(unit_str, product_name=""):
    u    = (unit_str or "").strip().lower()
    name = (product_name or "").strip().lower()

    try:
        # Standalone "кг" = price IS per kilogram
        if u == 'кг':
            return 1000, 'g'

        # If unit has a numeric measure, use it directly
        if u:
            result = _extract_measure(u)
            if result:
                return result

        # Piece unit: standalone "бр" or "бр." (no number)
        # → try to extract weight/volume from product name for meaningful per-kg/l comparison
        is_bare_piece = bool(re.fullmatch(r'бр\.?', u)) if u else False
        if is_bare_piece and name:
            result = _extract_measure(name)
            if result:
                return result  # e.g. "(150 г)" in name → лв./кг instead of лв./бр

        # Pieces with explicit count "N бр"
        if u:
            m = re.search(r'([\d]+[.,]?[\d]*)\s*бр', u)
            if m:
                return float(m.group(1).replace(',', '.')), 'pcs'

        # Fallback: try product name (for null/empty unit fields)
        s = name if not u else u
        if not s:
            return None, None

        result = _extract_measure(s)
        if result:
            return result

        m = re.search(r'([\d]+[.,]?[\d]*)\s*бр', s)
        if m:
            return float(m.group(1).replace(',', '.')), 'pcs'

        if 'бр' in s:
            return 1, 'pcs'

    except (ValueError, AttributeError):
        pass
    return None, None


def calc_norm(price, qty, base):
    if not (price and qty and qty > 0):
        return None, None
    if base == 'g':
        return round((price / qty) * 1000, 2), 'лв./кг'
    if base == 'ml':
        return round((price / qty) * 1000, 2), 'лв./л'
    if base == 'pcs':
        return round(price / qty, 2), 'лв./бр'
    return None, None


# ── Step 2: Classifier ────────────────────────────────────────────────────────

# Strips secondary ingredients after "с" / "със" so they don't trigger category rules.
# Example: "Крекери с кашкавал" → "Крекери"  (не Кашкавал)
#          "Die Maus хляб с извара" → "Die Maus хляб"  (не Сметана/Извара)
# The AI second-pass then receives the FULL original name and classifies correctly.
_WITH_RE = re.compile(r'^(.{3,}?)\s+с(?:ъс)?\s+\S', re.IGNORECASE)

def _primary_name(name: str) -> str:
    """Return the part of the product name before the first 'с/със + ingredient'."""
    m = _WITH_RE.match(name)
    return m.group(1).strip() if m else name


def classify(name):
    # Classify on the primary name only (strips secondary "с/със" ingredients)
    nl = _primary_name(name).lower()
    for cat, sub, incs, excs in RULES:
        if any(re.search(e, nl) for e in excs):
            continue
        if any(re.search(i, nl) for i in incs):
            return cat, sub
    return "Некласифицирани", "Некласифицирани"


# ── Step 3: Enrich ────────────────────────────────────────────────────────────
def enrich(items, overrides):
    out = []
    for item in items:
        name  = item.get('product_name', '')
        price = item.get('promo_price') or item.get('regular_price')
        reg   = item.get('regular_price')

        qty, base       = parse_unit(item.get('unit'), name)
        norm, norm_unit = calc_norm(price, qty, base)

        if name in overrides:
            ov  = overrides[name]
            cat = ov.get('category', 'Некласифицирани')
            sub = ov.get('subcategory', 'Некласифицирани')
        else:
            cat, sub = classify(name)

        discount = round(1 - price / reg, 4) if (price and reg and reg > price and reg > 0) else None

        out.append({
            **item,
            'category': cat, 'subcategory': sub,
            'parsed_qty': qty, 'base_unit': base,
            'norm_price': norm, 'norm_unit': norm_unit,
            'price': price, 'discount': discount,
        })
    return out


# ── Excel Helpers ─────────────────────────────────────────────────────────────
def title_row(ws, row, text, ncols, subtitle=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = SUBTITLE_FONT if subtitle else TITLE_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 22 if subtitle else 28


def header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.border    = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def cat_header(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text.upper())
    c.font      = CATROW_FONT
    c.fill      = CAT_FILL
    c.border    = THICK_BORDER
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 20


def subcat_header(ws, row, text, count, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = SUBCAT_FONT
    c.fill      = SUBCAT_FILL
    c.border    = BORDER
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    n = ws.cell(row=row, column=3, value=f"({count} продукта)")
    n.font  = NOTE_FONT
    n.fill  = SUBCAT_FILL
    n.border = BORDER
    for col in range(4, ncols + 1):
        ws.cell(row=row, column=col).fill   = SUBCAT_FILL
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 18


def ordered_groups(groups):
    result, seen = [], set()
    for cat in CAT_ORDER:
        for (c, s) in sorted(k for k in groups if k[0] == cat):
            if (c, s) not in seen:
                seen.add((c, s))
                result.append(((c, s), groups[(c, s)]))
    for key in sorted(k for k in groups if k not in seen):
        result.append((key, groups[key]))
    return result


def sort_key(cat):
    try:
        return CAT_ORDER.index(cat[0]), cat[1]
    except ValueError:
        return 99, cat[1]


# ── Sheet 1: Най-евтини по категория ─────────────────────────────────────────
def build_sheet1(wb, groups):
    ws = wb.create_sheet("Най-евтини по категория")
    ws.sheet_properties.tabColor = "00B050"

    COLS   = ["Категория", "Подкатегория", "#", "Продукт", "Магазин", "Канал",
              "Промо цена", "Опаковка", "Цена кг/л/бр", "Единица", "Отстъпка", "URL"]
    WIDTHS = [18, 22, 4, 50, 20, 10, 14, 14, 14, 10, 11, 45]
    FMTS   = [None]*6 + [MONEY_FMT, None, NORM_FMT, None, PCT_FMT, None]
    N      = len(COLS)

    title_row(ws, 1, "Най-евтини продукти по категория — Промоции 26.03–01.04.2026", N)
    title_row(ws, 2,
              "Цените са нормализирани (лв./кг, лв./л, лв./бр) за честно сравнение между "
              "различни размери опаковки. Показани са до 5 най-евтини варианта на подкатегория.", N, True)
    header_row(ws, 4, COLS)
    ws.freeze_panes = "A5"
    set_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 6

    row, prev_cat = 5, None
    for (cat, sub), items in ordered_groups(groups):
        top5 = sorted(items, key=lambda x: (x['norm_price'] is None, x['norm_price'] or 1e9))[:5]
        total = len(items)

        if cat != prev_cat:
            cat_header(ws, row, cat, N)
            row += 1
            prev_cat = cat

        subcat_header(ws, row, sub, total, N)
        row += 1

        for rank, it in enumerate(top5, 1):
            fill = GREEN if rank == 1 else (BLUE_LIGHT if rank == 2 else (GRAY if rank % 2 == 1 else None))
            url  = it.get('source_url') or ''
            vals = ["", sub, rank, it['product_name'], it['source_store'],
                    it['source_channel'], it['price'], it.get('unit') or '',
                    it['norm_price'], it['norm_unit'] or '', it['discount'], url]
            for col, (v, fmt) in enumerate(zip(vals, FMTS), 1):
                c = ws.cell(row=row, column=col, value=v)
                if col == N and url:
                    c.font      = URL_FONT
                    c.hyperlink = url
                else:
                    c.font = GREEN_FONT if (rank == 1 and col == 9) else (BOLD_FONT if rank == 1 else DATA_FONT)
                c.border = BORDER
                if fill:
                    c.fill = fill
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(wrap_text=(col == 4))
            ws.row_dimensions[row].height = 15
            row += 1

        ws.row_dimensions[row].height = 5
        row += 1


# ── Sheet 2: All by category ─────────────────────────────────────────────────
def build_sheet1b(wb, enriched):
    ws = wb.create_sheet("All by category")
    ws.sheet_properties.tabColor = "00B0F0"

    COLS   = ["Категория", "Подкатегория", "#", "Продукт", "Магазин", "Канал",
              "Промо цена", "Опаковка", "Цена кг/л/бр", "Единица", "Отстъпка", "URL"]
    WIDTHS = [18, 22, 4, 50, 20, 10, 14, 14, 14, 10, 11, 45]
    FMTS   = [None]*6 + [MONEY_FMT, None, NORM_FMT, None, PCT_FMT, None]
    N      = len(COLS)

    title_row(ws, 1, "Всички продукти по категория — Промоции", N)
    title_row(ws, 2,
              "Всички продукти, групирани по категория и подкатегория, наредени по нормализирана цена.", N, True)
    header_row(ws, 4, COLS)
    ws.freeze_panes = "A5"
    set_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 6

    all_groups = defaultdict(list)
    for it in enriched:
        all_groups[(it['category'], it['subcategory'])].append(it)

    row, prev_cat = 5, None
    for (cat, sub), items in ordered_groups(all_groups):
        sorted_items = sorted(items, key=lambda x: (x['norm_price'] is None, x['norm_price'] or 1e9))
        total = len(sorted_items)

        if cat != prev_cat:
            cat_header(ws, row, cat, N)
            row += 1
            prev_cat = cat

        subcat_header(ws, row, sub, total, N)
        row += 1

        for rank, it in enumerate(sorted_items, 1):
            fill = GREEN if rank == 1 else (BLUE_LIGHT if rank == 2 else (GRAY if rank % 2 == 1 else None))
            url  = it.get('source_url') or ''
            vals = ["", sub, rank, it['product_name'], it['source_store'],
                    it['source_channel'], it['price'], it.get('unit') or '',
                    it['norm_price'], it['norm_unit'] or '', it['discount'], url]
            for col, (v, fmt) in enumerate(zip(vals, FMTS), 1):
                c = ws.cell(row=row, column=col, value=v)
                if col == N and url:
                    c.font      = URL_FONT
                    c.hyperlink = url
                else:
                    c.font = GREEN_FONT if (rank == 1 and col == 9) else (BOLD_FONT if rank == 1 else DATA_FONT)
                c.border = BORDER
                if fill:
                    c.fill = fill
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(wrap_text=(col == 4))
            ws.row_dimensions[row].height = 15
            row += 1

        ws.row_dimensions[row].height = 5
        row += 1


# ── Sheet 3: Обобщение ────────────────────────────────────────────────────────
def build_sheet2(wb, groups):
    ws = wb.create_sheet("Обобщение")
    ws.sheet_properties.tabColor = "2E75B6"

    COLS   = ["Категория", "Подкатегория", "Най-евтин продукт", "Магазин", "Канал",
              "Промо цена", "Опаковка", "Норм. цена", "Единица", "# Алтернативи", "URL"]
    WIDTHS = [18, 22, 50, 20, 10, 14, 14, 16, 10, 14, 45]
    FMTS   = [None]*5 + [MONEY_FMT, None, NORM_FMT, None, None, None]
    N      = len(COLS)

    title_row(ws, 1, "Най-евтиният продукт за всяка подкатегория", N)
    title_row(ws, 2, "Бърза справка — само един ред на подкатегория, с найевтиния наличен продукт.", N, True)
    header_row(ws, 4, COLS)
    ws.freeze_panes = "A5"
    set_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 6

    row, prev_cat, even = 5, None, False
    for (cat, sub), items in ordered_groups(groups):
        best  = sorted(items, key=lambda x: (x['norm_price'] is None, x['norm_price'] or 1e9))[0]
        count = len(items)

        if cat != prev_cat:
            cat_header(ws, row, cat, N)
            row += 1
            prev_cat = cat
            even = False

        fill = GRAY if even else None
        even = not even
        url  = best.get('source_url') or ''
        vals = [cat, sub, best['product_name'], best['source_store'], best['source_channel'],
                best['price'], best.get('unit') or '', best['norm_price'],
                best['norm_unit'] or '', count - 1, url]
        for col, (v, fmt) in enumerate(zip(vals, FMTS), 1):
            c = ws.cell(row=row, column=col, value=v)
            if col == N and url:
                c.font      = URL_FONT
                c.hyperlink = url
            elif col == 8:
                c.font = GREEN_FONT
            else:
                c.font = DATA_FONT
            c.border = BORDER
            if col == 8:
                c.fill = GREEN
            elif fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(wrap_text=(col == 3))
        ws.row_dimensions[row].height = 15
        row += 1


# ── Union-Find (Sheet 3) ──────────────────────────────────────────────────────
class UnionFind:
    def __init__(self, n):
        self.p = list(range(n))

    def find(self, x):
        while self.p[x] != x:
            self.p[x] = self.p[self.p[x]]
            x = self.p[x]
        return x

    def union(self, x, y):
        px, py = self.find(x), self.find(y)
        if px != py:
            self.p[py] = px


_STOP = {
    'различни', 'видове', 'вид', 'бр', 'кг', 'гр', 'мл', 'от', 'за', 'с', 'и', 'на',
    'в', 'до', 'по', 'при', 'към', 'без', 'или', 'лв', 'избрани', 'нашата',
    'пекарна', 'свежата', 'витрина', 'цена', 'the', 'and',
}


def keywords(name):
    words = re.findall(r'[а-яА-Яa-zA-Z]{3,}', name.lower())
    return {w for w in words if w not in _STOP}


def auto_clusters(items):
    n = len(items)
    if n < 2:
        return []
    uf  = UnionFind(n)
    kws = [keywords(it['product_name']) for it in items]
    for i in range(n):
        for j in range(i + 1, n):
            if items[i]['source_store'] != items[j]['source_store']:
                if len(kws[i] & kws[j]) >= 2:
                    uf.union(i, j)
    buckets = defaultdict(list)
    for i in range(n):
        buckets[uf.find(i)].append(items[i])
    return [cl for cl in buckets.values()
            if len({it['source_store'] for it in cl}) >= 2]


# ── Sheet 3: Сравнение (авто) ─────────────────────────────────────────────────
def build_sheet3(wb, enriched):
    ws = wb.create_sheet("Сравнение (авто)")
    ws.sheet_properties.tabColor = "7030A0"

    S = 5  # max stores per row
    BASE_COLS = ["Категория", "Подкатегория", "Общи думи"]
    STORE_COLS = []
    for i in range(1, S + 1):
        STORE_COLS += [f"Продукт {i}", f"Магазин {i}", f"Цена {i}", f"Норм. цена {i}"]
    ALL = BASE_COLS + STORE_COLS
    N   = len(ALL)
    FMTS = [None, None, None] + [None, None, MONEY_FMT, NORM_FMT] * S

    title_row(ws, 1, "Автоматично сравнение — съвпадения по ключови думи (Union-Find)", N)
    title_row(ws, 2,
              "Алгоритъмът открива продукти от различни магазини в една подкатегория "
              "с ≥2 съвпадащи ключови думи. Подредени по брой магазини (повече = по-горе).", N, True)
    header_row(ws, 4, ALL)
    ws.freeze_panes = "A5"
    set_widths(ws, [18, 22, 28] + [36, 16, 12, 12] * S)
    ws.row_dimensions[3].height = 6

    by_sub = defaultdict(list)
    for it in enriched:
        if it['category'] != 'Некласифицирани':
            by_sub[(it['category'], it['subcategory'])].append(it)

    row, prev_cat, even, total = 5, None, False, 0
    for key in sorted(by_sub, key=sort_key):
        cat, sub = key
        clusters = auto_clusters(by_sub[key])
        if not clusters:
            continue
        clusters.sort(key=lambda cl: len({x['source_store'] for x in cl}), reverse=True)

        if cat != prev_cat:
            cat_header(ws, row, cat, N)
            row += 1
            prev_cat = cat
            even = False

        for cl in clusters:
            cl.sort(key=lambda x: (x['norm_price'] is None, x['norm_price'] or 1e9))
            common = set.intersection(*[keywords(it['product_name']) for it in cl]) if cl else set()
            fill   = GRAY if even else None
            even   = not even

            vals = [cat, sub, ', '.join(sorted(common)[:6])]
            for it in cl[:S]:
                vals += [it['product_name'], it['source_store'],
                         it['price'], it['norm_price'] or '']
            while len(vals) < N:
                vals.append('')

            for col, (v, fmt) in enumerate(zip(vals[:N], FMTS[:N]), 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font   = DATA_FONT
                c.border = BORDER
                if fill:
                    c.fill = fill
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(wrap_text=(col in (3, 4, 8, 12, 16, 20)))
            ws.row_dimensions[row].height = 15
            row += 1
            total += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    f = ws.cell(row=row, column=1,
                value=f"Общо намерени групи: {total}  |  Метод: Union-Find, ≥2 ключови думи")
    f.font      = NOTE_FONT
    f.alignment = Alignment(horizontal='right')


# ── Valid categories derived from RULES ──────────────────────────────────────
_VALID_CATS: set[tuple] = set()
for _cat, _sub, _, _ in RULES:
    _VALID_CATS.add((_cat, _sub))

# Formatted list for prompts
_CAT_LIST_STR = '\n'.join(
    f"  {cat} → {sub}"
    for cat, sub in sorted(_VALID_CATS, key=lambda x: (CAT_ORDER.index(x[0]) if x[0] in CAT_ORDER else 99, x[1]))
)


# ── AI Classification of Unclassified Items ───────────────────────────────────
def ai_classify_batch(cfg, key, batch):
    """
    Ask GPT-4o to classify a batch of unclassified product names.
    Returns dict: {product_name: (category, subcategory)} for confident matches only.
    Items where GPT-4o is not confident are NOT included (caller treats them as unclassified).
    """
    try:
        from openai import AzureOpenAI
    except ImportError:
        return {}

    client = AzureOpenAI(
        azure_endpoint=cfg['endpoint'],
        api_key=key,
        api_version=cfg['api_version'],
        timeout=cfg.get('timeout_seconds', 60),
    )

    items_str = '\n'.join(f'{i+1}. {it["product_name"]}' for i, it in enumerate(batch))

    prompt = (
        "Ти класифицираш продукти от български супермаркети.\n\n"
        f"ЕДИНСТВЕНО ВАЛИДНИ КАТЕГОРИИ (само от този списък — никакви други!):\n"
        f"{_CAT_LIST_STR}\n\n"
        "СТРОГО ПРАВИЛО: Ако не си АБСОЛЮТНО СИГУРЕН в коя категория и подкатегория "
        "принадлежи продуктът — върни null и за двете полета. "
        "НЕ ИЗМИСЛЯЙ категории извън горния списък! "
        "По-добре null, отколкото грешна категория.\n\n"
        f"Продукти:\n{items_str}\n\n"
        'Върни САМО JSON:\n'
        '{"results": [{"index": 1, "category": "Млечни продукти", "subcategory": "Кисело мляко"}, '
        '{"index": 2, "category": null, "subcategory": null}]}'
    )

    try:
        resp = client.chat.completions.create(
            model=cfg['deployment_name'],
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            max_tokens=cfg.get('max_tokens', 2000),
            temperature=0.0,
        )
        data = json.loads(resp.choices[0].message.content)
    except Exception as e:
        print(f"    Грешка: {e}", flush=True)
        return {}

    result = {}
    for r in data.get('results', []):
        idx = r.get('index', 0) - 1
        if not (0 <= idx < len(batch)):
            continue
        cat = r.get('category')
        sub = r.get('subcategory')
        name = batch[idx]['product_name']
        # Hard validation: only accept if exactly in _VALID_CATS
        if cat and sub and (cat, sub) in _VALID_CATS:
            result[name] = (cat, sub)
        # else: leave out — will remain Некласифицирани → goes to За преглед
    return result


def ai_classify_unclassified(enriched, cfg, key):
    """
    Run a second classification pass using GPT-4o on all Некласифицирани items.
    Mutates enriched in place: updates category/subcategory for confidently classified items.
    Returns count of newly classified items.
    """
    unclassified = [it for it in enriched if it['category'] == 'Некласифицирани']
    if not unclassified:
        return 0

    batch_size = cfg.get('max_items_per_call', 40)
    delay      = cfg.get('delay_between_calls_seconds', 0.5)
    total_new  = 0
    batches    = [unclassified[i:i+batch_size] for i in range(0, len(unclassified), batch_size)]

    print(f"  ИИ класификация: {len(unclassified)} некласифицирани → {len(batches)} заявки...",
          flush=True)

    for b_idx, batch in enumerate(batches, 1):
        print(f"    Партида {b_idx}/{len(batches)} ({len(batch)} продукта)...", flush=True)
        classified_map = ai_classify_batch(cfg, key, batch)

        # Apply back to enriched
        name_to_item = {it['product_name']: it for it in batch}
        for name, (cat, sub) in classified_map.items():
            if name in name_to_item:
                name_to_item[name]['category']    = cat
                name_to_item[name]['subcategory'] = sub
                total_new += 1

        if b_idx < len(batches):
            time.sleep(delay)

    return total_new


# ── Azure OpenAI ──────────────────────────────────────────────────────────────
def load_azure():
    try:
        cfg = json.loads(CONFIG_PATH.read_text(encoding='utf-8'))
        sec = json.loads(SECRETS_PATH.read_text(encoding='utf-8'))
        key = sec.get('api_key', '')
        if 'ВАШИЯТ' in key or not key:
            print("  Предупреждение: попълнете api_key в azure_secrets.json", flush=True)
            return cfg, None
        return cfg, key
    except Exception as e:
        print(f"  Грешка при Azure конфиг: {e}", flush=True)
        return {}, None


def ai_match(cfg, key, subcat, items):
    try:
        from openai import AzureOpenAI
    except ImportError:
        return []
    client = AzureOpenAI(
        azure_endpoint=cfg['endpoint'],
        api_key=key,
        api_version=cfg['api_version'],
        timeout=cfg.get('timeout_seconds', 60),
    )
    lines = [
        f"{i+1}. [{it['source_store']}] {it['product_name']} — "
        f"{it['price']:.2f} лв. ({it.get('unit') or 'н/п'})"
        for i, it in enumerate(items)
    ]
    prompt = (
        f'Категория: "{subcat}"\n\n'
        f'Намери ЕКВИВАЛЕНТНИ продукти (същият тип, от различни магазини).\n'
        f'ПРАВИЛО: НЕ ИЗМИСЛЯЙ — само ако си сигурен 100%!\n'
        f'Само различни магазини могат да се групират.\n\n'
        f'Продукти:\n' + '\n'.join(lines) + '\n\n'
        f'Върни САМО JSON: {{"groups": [[1,5],[3,8]]}} или {{"groups": []}}'
    )
    try:
        resp = client.chat.completions.create(
            model=cfg['deployment_name'],
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            max_tokens=cfg.get('max_tokens', 2000),
            temperature=cfg.get('temperature', 0.0),
        )
        return json.loads(resp.choices[0].message.content).get('groups', [])
    except Exception as e:
        print(f"  OpenAI грешка: {e}", flush=True)
        return []


# ── Sheet 4: Сравнение (ИИ) ───────────────────────────────────────────────────
def build_sheet4(wb, enriched, cfg, key):
    ws = wb.create_sheet("Сравнение (ИИ)")
    ws.sheet_properties.tabColor = "C00000"

    S = 5
    BASE_COLS  = ["Категория", "Подкатегория", "Група"]
    STORE_COLS = []
    for i in range(1, S + 1):
        STORE_COLS += [f"Продукт {i}", f"Магазин {i}", f"Цена {i}", f"Норм. цена {i}"]
    ALL  = BASE_COLS + STORE_COLS
    N    = len(ALL)
    FMTS = [None, None, None] + [None, None, MONEY_FMT, NORM_FMT] * S

    title_row(ws, 1, "Сравнение с изкуствен интелект — Azure OpenAI GPT-4o", N)
    title_row(ws, 2,
              "ИИ открива семантично еквивалентни продукти между магазините. "
              "Инструкция: да не се измислят съответствия — само при 100% сигурност.", N, True)
    header_row(ws, 4, ALL)
    ws.freeze_panes = "A5"
    set_widths(ws, [18, 22, 8] + [36, 16, 12, 12] * S)
    ws.row_dimensions[3].height = 6

    if not key:
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=N)
        c = ws.cell(row=5, column=1,
                    value="API ключът не е попълнен в azure_secrets.json. "
                          "Попълнете го и стартирайте скрипта отново.")
        c.font      = Font(name='Arial', bold=True, size=11, color='C00000')
        c.alignment = Alignment(horizontal='center')
        return

    by_sub = defaultdict(list)
    for it in enriched:
        if it['category'] != 'Некласифицирани':
            by_sub[(it['category'], it['subcategory'])].append(it)

    row, prev_cat, even = 5, None, False
    grp_num, total = 0, 0
    delay      = cfg.get('delay_between_calls_seconds', 0.5)
    max_items  = cfg.get('max_items_per_call', 40)

    for key_sub in sorted(by_sub, key=sort_key):
        cat, sub = key_sub
        pool     = by_sub[key_sub]
        stores   = {it['source_store'] for it in pool}
        if len(stores) < 2:
            continue

        # Sample diverse across stores
        store_q = defaultdict(list)
        for it in pool:
            store_q[it['source_store']].append(it)
        sampled = []
        while len(sampled) < max_items and any(store_q.values()):
            for st in list(store_q):
                if store_q[st] and len(sampled) < max_items:
                    sampled.append(store_q[st].pop(0))

        print(f"  ИИ: {sub} ({len(sampled)} продукта)...", flush=True)
        ai_groups = ai_match(cfg, key, sub, sampled)
        if not ai_groups:
            time.sleep(delay)
            continue

        if cat != prev_cat:
            cat_header(ws, row, cat, N)
            row += 1
            prev_cat = cat
            even = False

        for grp in ai_groups:
            try:
                grp_items = [sampled[i - 1] for i in grp if 1 <= i <= len(sampled)]
            except (IndexError, TypeError):
                continue
            if len({it['source_store'] for it in grp_items}) < 2:
                continue

            grp_num += 1
            grp_items.sort(key=lambda x: (x['norm_price'] is None, x['norm_price'] or 1e9))
            fill = GRAY if even else None
            even = not even

            vals = [cat, sub, grp_num]
            for it in grp_items[:S]:
                vals += [it['product_name'], it['source_store'],
                         it['price'], it['norm_price'] or '']
            while len(vals) < N:
                vals.append('')

            for col, (v, fmt) in enumerate(zip(vals[:N], FMTS[:N]), 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font   = DATA_FONT
                c.border = BORDER
                if fill:
                    c.fill = fill
                if fmt:
                    c.number_format = fmt
                c.alignment = Alignment(wrap_text=(col in (4, 8, 12, 16, 20)))
            ws.row_dimensions[row].height = 15
            row += 1
            total += 1

        time.sleep(delay)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    f = ws.cell(row=row, column=1,
                value=f"Общо ИИ-групи: {total}  |  Модел: GPT-4o (Azure)  |  "
                      f"Температура: {cfg.get('temperature', 0.0)}")
    f.font      = NOTE_FONT
    f.alignment = Alignment(horizontal='right')


# ── Sheet 5: Всички продукти ──────────────────────────────────────────────────
def build_sheet5(wb, enriched):
    ws = wb.create_sheet("Всички продукти")
    ws.sheet_properties.tabColor = "1F4E79"

    COLS   = ["#", "Категория", "Подкатегория", "Продукт", "Магазин", "Канал",
              "Редовна цена", "Промо цена", "Отстъпка", "Спестяване",
              "Опаковка", "Норм. цена", "Единица", "Промо период", "URL"]
    WIDTHS = [5, 18, 22, 50, 18, 10, 14, 14, 10, 12, 14, 14, 10, 18, 55]
    FMTS   = [None]*6 + [MONEY_FMT, MONEY_FMT, PCT_FMT, MONEY_FMT,
                         None, NORM_FMT, None, None, None]
    N      = len(COLS)

    title_row(ws, 1, f"Всички промоционални продукти — {len(enriched)} записа", N)
    header_row(ws, 2, COLS)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(N)}2"
    set_widths(ws, WIDTHS)

    for i, it in enumerate(enriched, 1):
        r   = i + 2
        reg = it.get('regular_price')
        p   = it['price']
        sav = round(reg - p, 2) if (reg and p and reg > p) else None
        url = it.get('source_url') or ''

        vals = [i, it['category'], it['subcategory'], it['product_name'],
                it['source_store'], it['source_channel'], reg, p,
                it['discount'], sav, it.get('unit') or '',
                it['norm_price'], it['norm_unit'] or '',
                it.get('promo_period') or '', url]
        fill = GRAY if i % 2 == 0 else None
        for col, (v, fmt) in enumerate(zip(vals, FMTS), 1):
            c = ws.cell(row=r, column=col, value=v)
            if col == N and url:
                c.font      = URL_FONT
                c.hyperlink = url
            else:
                c.font = DATA_FONT
            c.border = BORDER
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(wrap_text=(col in (4, N)))
        ws.row_dimensions[r].height = 15

    ws.row_dimensions[1].height = 26


# ── Sheet 6: За преглед ───────────────────────────────────────────────────────
def build_sheet6(wb, enriched):
    ws = wb.create_sheet("За преглед")
    ws.sheet_properties.tabColor = "FFC000"

    review = [it for it in enriched
              if it['category'] == 'Некласифицирани' or it['norm_price'] is None]

    COLS   = ["#", "Продукт", "Магазин", "Промо цена", "Опаковка (сурово)",
              "Причина", "Текуща категория",
              "Правилна категория", "Правилна подкатегория", "Бележка", "URL"]
    WIDTHS = [4, 50, 16, 12, 16, 28, 20, 22, 22, 30, 45]
    N      = len(COLS)

    title_row(ws, 1, f"Продукти за преглед — {len(review)} записа", N)
    title_row(ws, 2,
              "Попълнете 'Правилна категория' и 'Правилна подкатегория', след което добавете "
              "корекциите в manual_overrides.json → ключ 'overrides'.", N, True)
    ws.row_dimensions[2].height = 24
    header_row(ws, 4, COLS)
    ws.freeze_panes = "A5"
    set_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 6

    dv = DataValidation(
        type="list",
        formula1="Настройки!$A$1:$A$17",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Невалидна категория",
        error="Изберете от падащия списък.",
    )
    ws.add_data_validation(dv)

    for i, it in enumerate(review, 1):
        r   = i + 4
        unc = it['category'] == 'Некласифицирани'
        nop = it['norm_price'] is None
        if unc and nop:
            reason = "Некласифициран + без единица"
        elif unc:
            reason = "Некласифициран"
        else:
            reason = "Без разпозната единица"

        url  = it.get('source_url') or ''
        vals = [i, it['product_name'], it['source_store'], it['price'],
                it.get('unit') or '(няма)', reason, it['category'],
                '', '', '', url]
        fill = GRAY if i % 2 == 0 else None
        FMTS = [None, None, None, MONEY_FMT, None, None, None, None, None, None, None]
        for col, (v, fmt) in enumerate(zip(vals, FMTS), 1):
            c = ws.cell(row=r, column=col, value=v)
            if col == N and url:
                c.font      = URL_FONT
                c.hyperlink = url
            elif col == 6:
                c.font = BOLD_FONT
                c.fill = YELLOW
            elif col in (8, 9, 10):
                c.font = DATA_FONT
                c.fill = REVIEW_FILL
            else:
                c.font = DATA_FONT
                if fill:
                    c.fill = fill
            c.border = BORDER
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(wrap_text=(col in (2, 10)))
        dv.add(f"H{r}")
        ws.row_dimensions[r].height = 15



# ── Hidden Settings Sheet ─────────────────────────────────────────────────────
def build_settings(wb):
    ws = wb.create_sheet("Настройки")
    ws.sheet_state = 'hidden'
    for i, cat in enumerate(CAT_ORDER, 1):
        ws.cell(row=i, column=1, value=cat)


# ── Main ──────────────────────────────────────────────────────────────────────
def main(args=None):
    print(f"Зареждам {MASTER_PATH.name}...", flush=True)
    raw = json.loads(MASTER_PATH.read_text(encoding='utf-8'))
    print(f"  {len(raw)} записа", flush=True)

    overrides = {}
    if OVERRIDES_PATH.exists():
        try:
            data      = json.loads(OVERRIDES_PATH.read_text(encoding='utf-8'))
            overrides = data.get('overrides', {})
            if overrides:
                print(f"  Ръчни корекции: {len(overrides)}", flush=True)
        except Exception as e:
            print(f"  Предупреждение: {OVERRIDES_PATH.name}: {e}", flush=True)

    print("Обогатявам данните...", flush=True)
    enriched = enrich(raw, overrides)
    classified = sum(1 for it in enriched if it['category'] != 'Некласифицирани')
    with_norm  = sum(1 for it in enriched if it['norm_price'] is not None)
    n = len(enriched)
    print(f"  Класифицирани: {classified}/{n} ({100*classified//n}%)", flush=True)
    print(f"  С нормализирана цена: {with_norm}/{n} ({100*with_norm//n}%)", flush=True)

    # Groups for sheets 1 & 2 (classified + with norm_price)
    groups: dict = defaultdict(list)
    for it in enriched:
        if it['norm_price'] is not None and it['category'] != 'Некласифицирани':
            groups[(it['category'], it['subcategory'])].append(it)
    print(f"  Подкатегории с нормализирана цена: {len(groups)}", flush=True)

    print("Зареждам Azure конфигурация...", flush=True)
    az_cfg, az_key = load_azure()

    # AI second-pass classification for unclassified items
    if az_key:
        newly = ai_classify_unclassified(enriched, az_cfg, az_key)
        if newly:
            classified = sum(1 for it in enriched if it['category'] != 'Некласифицирани')
            print(f"  След ИИ — класифицирани: {classified}/{n} ({100*classified//n}%)",
                  flush=True)
            # Rebuild groups with updated classifications
            groups = defaultdict(list)
            for it in enriched:
                if it['norm_price'] is not None and it['category'] != 'Некласифицирани':
                    groups[(it['category'], it['subcategory'])].append(it)
            print(f"  Подкатегории след ИИ: {len(groups)}", flush=True)

    wb = Workbook()
    wb.remove(wb.active)

    print("  Лист 1: Най-евтини по категория...", flush=True)
    build_sheet1(wb, groups)

    print("  Лист 2: All by category...", flush=True)
    build_sheet1b(wb, enriched)

    print("  Лист 3: Обобщение...", flush=True)
    build_sheet2(wb, groups)

    print("  Лист 4: Сравнение (авто)...", flush=True)
    build_sheet3(wb, enriched)

    print("  Лист 5: Сравнение (ИИ)...", flush=True)
    build_sheet4(wb, enriched, az_cfg, az_key)

    print("  Лист 6: Всички продукти...", flush=True)
    build_sheet5(wb, enriched)

    print("  Лист 7: За преглед...", flush=True)
    build_sheet6(wb, enriched)

    build_settings(wb)

    save_path = OUTPUT_PATH
    try:
        wb.save(save_path)
    except PermissionError:
        save_path = OUTPUT_PATH.with_stem(OUTPUT_PATH.stem + "_new")
        wb.save(save_path)
        print(f"\nПредупреждение: {OUTPUT_PATH.name} е отворен в Excel — записан като {save_path.name}",
              flush=True)
        print("Затворете оригиналния файл и преименувайте.", flush=True)
    size_kb = save_path.stat().st_size // 1024
    print(f"\nГотово! {save_path.name}  ({size_kb} KB)", flush=True)
    unclass = n - classified
    no_unit = n - with_norm
    print(f"  Некласифицирани: {unclass}  |  Без единица: {no_unit}", flush=True)

    # ── English version (--english flag) ──────────────────────────────────────
    if args and args.english:
        if not az_key:
            print("\nПредупреждение: Azure OpenAI ключът не е наличен — "
                  "английската версия не е генерирана.", flush=True)
        else:
            print("\nГенерирам английска версия...", flush=True)
            import translator as _tr
            en_wb   = load_workbook(str(save_path))
            _tr.translate_workbook(en_wb, az_cfg, az_key, batch_size=50, verbose=True)
            en_path = save_path.with_stem(save_path.stem + "_en")
            try:
                en_wb.save(str(en_path))
            except PermissionError:
                en_path = en_path.with_stem(en_path.stem + "_new")
                en_wb.save(str(en_path))
                print(f"  Файлът е отворен в Excel — записан като {en_path.name}", flush=True)
            en_kb = en_path.stat().st_size // 1024
            print(f"  English: {en_path.name}  ({en_kb} KB)", flush=True)


if __name__ == '__main__':
    import argparse as _ap
    _parser = _ap.ArgumentParser(description='Generate Bulgarian promo price report.')
    _parser.add_argument(
        '--english', action='store_true',
        help='Also generate an English copy (<name>_en.xlsx) via GPT-4o translation.',
    )
    _args = _parser.parse_args()
    main(_args)
