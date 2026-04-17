#!/usr/bin/env python3
"""
Bulgarian Grocery Promo Prices — XLSX Generator
Reads bulgarian_promo_prices_merged.json and produces a formatted Excel file.
"""

import json
import re
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
INPUT_PATH = BASE_DIR / "bulgarian_promo_prices_merged.json"

# ── Styling constants ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT     = Font(name='Arial', size=10)
TITLE_FONT    = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='1F4E79')
NOTE_FONT     = Font(name='Arial', italic=True, size=10, color='808080')
GREEN_FONT    = Font(name='Arial', bold=True, size=10, color='006100')
RED_FONT      = Font(name='Arial', bold=True, size=10, color='9C0006')

MONEY_FORMAT = '#,##0.00" лв."'
PCT_FORMAT   = '0%'

GREEN_FILL  = PatternFill('solid', fgColor='E2EFDA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
RED_FILL    = PatternFill('solid', fgColor='FCE4EC')
GRAY_FILL   = PatternFill('solid', fgColor='F2F2F2')
BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))


def style_header(ws, row, num_columns):
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_data_row(ws, row, num_columns, alternating=False):
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = BORDER
        if alternating:
            cell.fill = GRAY_FILL


# ── Data quality checks ───────────────────────────────────────────────────────
def run_quality_checks(data):
    print("\n=== DATA QUALITY CHECKS ===")

    # 2.1 Basic counts
    print(f"\nTotal items: {len(data)}")
    by_store = Counter(d['source_store'] for d in data)
    print("\nBy store:")
    for k, v in by_store.most_common():
        print(f"  {k}: {v}")
    by_channel = Counter(d['source_channel'] for d in data)
    print("\nBy channel:")
    for k, v in by_channel.most_common():
        print(f"  {k}: {v}")
    by_combo = Counter((d['source_store'], d['source_channel']) for d in data)
    print("\nBy (store, channel):")
    for k, v in by_combo.most_common():
        print(f"  {k[0]} / {k[1]}: {v}")

    # 2.2 Price validation
    no_price = sum(1 for d in data if not d.get('promo_price') and not d.get('regular_price'))
    no_name  = sum(1 for d in data if not d.get('product_name'))
    has_both = sum(1 for d in data if d.get('promo_price') and d.get('regular_price'))
    suspect  = sum(1 for d in data
                   if d.get('promo_price') and d.get('regular_price')
                   and d['promo_price'] > d['regular_price'])
    print(f"\nNo price (both null): {no_price}")
    print(f"No name: {no_name}")
    print(f"Has both prices: {has_both}")
    print(f"Suspect (promo > regular): {suspect}")

    # 2.3 Category coverage
    cat_counter = Counter(d.get('product_category') for d in data)
    print(f"\nNull category: {cat_counter[None]}")
    print(f"Items with category: {len(data) - cat_counter[None]}")

    # 2.5 Price stats per source
    print("\nPrice stats per (store, channel):")
    by_source = defaultdict(list)
    for d in data:
        p = d.get('promo_price')
        if p:
            by_source[(d['source_store'], d['source_channel'])].append(p)
    for key in sorted(by_source.keys()):
        vals = by_source[key]
        print(f"  {key[0]} / {key[1]}: n={len(vals)}, "
              f"min={min(vals):.2f}, max={max(vals):.2f}, "
              f"mean={statistics.mean(vals):.2f}, median={statistics.median(vals):.2f}")

    # 2.6 Top discounts
    discounts = []
    for d in data:
        reg = d.get('regular_price')
        promo = d.get('promo_price')
        if reg and promo and reg > 0:
            discounts.append((1 - promo / reg, d))
    discounts.sort(reverse=True, key=lambda x: x[0])
    print("\nTop 15 discounts:")
    for pct, d in discounts[:15]:
        print(f"  {pct:.0%}  {d['product_name'][:50]:<50}  "
              f"{d['source_store']:<20}  "
              f"{d['regular_price']:.2f} -> {d['promo_price']:.2f} lv.")


# ── Fuzzy matching ─────────────────────────────────────────────────────────────
_STOPWORDS = {'г', 'мл', 'кг', 'л', 'бр', 'бут', 'пакет', 'различни',
              'видове', 'клас', 'произход', 'пр', 'д', 'без', 'или', 'за',
              'до', 'при', 'на', 'от', 'по', 'и', 'в', 'с'}

def normalize_name(name):
    n = name.lower().strip()
    n = re.sub(r'\s+', ' ', n)
    n = re.sub(r'\d+\s*(г|мл|кг|л|бр)\s*$', '', n).strip()
    return n


def extract_keywords(name):
    n = name.lower().strip()
    n = re.sub(r'[/\-–—()]', ' ', n)
    n = re.sub(r'\d+', '', n)
    n = re.sub(r'\b(г|мл|кг|л|бр|бут|пакет|различни|видове|клас|произход|пр|д)\b', '', n)
    words = set(w for w in n.split() if len(w) > 2 and w not in _STOPWORDS)
    return words


def find_cross_store_matches(data):
    by_source = defaultdict(list)
    for item in data:
        key = (item['source_store'], item['source_channel'])
        by_source[key].append(item)

    sources = list(by_source.keys())
    matches = []

    for i in range(len(sources)):
        for j in range(i + 1, len(sources)):
            for item_a in by_source[sources[i]]:
                kw_a = extract_keywords(item_a['product_name'])
                if len(kw_a) < 2:
                    continue
                for item_b in by_source[sources[j]]:
                    kw_b = extract_keywords(item_b['product_name'])
                    overlap = kw_a & kw_b
                    if len(overlap) >= 2:
                        matches.append({
                            'overlap_count': len(overlap),
                            'overlap_words': overlap,
                            'item_a': item_a,
                            'item_b': item_b,
                        })

    matches.sort(key=lambda x: -x['overlap_count'])
    return matches


# ── Sheet 1: All Items ─────────────────────────────────────────────────────────
def build_all_items_sheet(wb, data):
    ws = wb.create_sheet("All Items")
    ws.sheet_properties.tabColor = '1F4E79'

    headers = ['#', 'Store', 'Channel', 'Product Name', 'Category',
               'Regular Price', 'Promo Price', 'Discount %', 'Savings',
               'Unit', 'Price/Unit', 'Promo Period', 'Source URL']
    col_widths = [5, 18, 10, 50, 22, 14, 14, 12, 12, 12, 14, 20, 55]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=col, value=h)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    style_header(ws, 1, len(headers))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:M1'

    for i, item in enumerate(data, 1):
        row = i + 1
        reg   = item.get('regular_price')
        promo = item.get('promo_price')
        disc  = (1 - promo / reg) if (reg and promo and reg > 0) else None
        sav   = (reg - promo) if (reg and promo) else None

        values = [
            i,
            item.get('source_store', ''),
            item.get('source_channel', ''),
            item.get('product_name', ''),
            item.get('product_category') or '',
            reg,
            promo,
            disc,
            sav,
            item.get('unit') or '',
            item.get('price_per_unit') or '',
            item.get('promo_period') or '',
            item.get('source_url', ''),
        ]

        alternating = (i % 2 == 0)
        style_data_row(ws, row, len(headers), alternating=alternating)

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)

        # Number formats
        for col in [6, 7, 9]:
            ws.cell(row=row, column=col).number_format = MONEY_FORMAT
        ws.cell(row=row, column=8).number_format = PCT_FORMAT

        # Conditional formatting on Discount % (col 8)
        if disc is not None:
            disc_cell = ws.cell(row=row, column=8)
            if disc >= 0.50:
                disc_cell.fill = GREEN_FILL
                disc_cell.font = GREEN_FONT
            elif disc >= 0.30:
                disc_cell.fill = YELLOW_FILL
                disc_cell.font = DATA_FONT

    # Update auto-filter range
    ws.auto_filter.ref = f'A1:M{len(data) + 1}'
    print(f"Sheet 1 'All Items': {len(data)} rows written.")


# ── Sheet 2: Cross-Store Comparison ───────────────────────────────────────────
def build_cross_store_sheet(wb, data):
    ws = wb.create_sheet("Cross-Store Comparison")
    ws.sheet_properties.tabColor = '2E75B6'

    # Title rows
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Cross-Store Price Comparison'
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:K2')
    ws['A2'] = ('Fuzzy keyword matching (≥2 shared words). '
                'Same product may have very different names across stores. '
                'Review carefully — false positives possible.')
    ws['A2'].font = NOTE_FONT

    # Blank spacer row 3
    headers = ['Product Names', 'Store 1', 'Ch.1', 'Price 1', 'URL 1',
               'Store 2', 'Ch.2', 'Price 2', 'URL 2', 'Price Diff', 'Cheaper At']
    col_widths = [50, 18, 10, 14, 40, 18, 10, 14, 40, 14, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=4, column=col, value=h)
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = w
    style_header(ws, 4, len(headers))
    ws.freeze_panes = 'A5'

    matches = find_cross_store_matches(data)
    print(f"\nFuzzy matching: {len(matches)} raw matches found.")

    # Auto-filter high-quality matches (overlap >= 3) and cap at 200
    good_matches = [m for m in matches if m['overlap_count'] >= 3]
    if len(good_matches) < 30:
        good_matches = matches[:100]  # fall back to top 100

    # Deduplicate: avoid showing the same pair of products twice
    seen_pairs = set()
    curated = []
    for m in good_matches:
        a = m['item_a']
        b = m['item_b']
        # Cheap key to prevent same pair reversed
        key = tuple(sorted([
            (a['source_store'], a['source_channel'], a['product_name'][:30]),
            (b['source_store'], b['source_channel'], b['product_name'][:30]),
        ]))
        if key not in seen_pairs:
            seen_pairs.add(key)
            curated.append(m)

    print(f"After dedup: {len(curated)} matches in Sheet 2.")

    for i, m in enumerate(curated):
        row = i + 5
        a, b = m['item_a'], m['item_b']
        price_a = a.get('promo_price') or a.get('regular_price')
        price_b = b.get('promo_price') or b.get('regular_price')

        if price_a is None or price_b is None:
            continue

        # Put cheaper item as "Store 1"
        if price_a > price_b:
            a, b = b, a
            price_a, price_b = price_b, price_a

        diff = abs(price_b - price_a)
        cheaper_at = f"{a['source_store']} ({a['source_channel']})"
        product_label = f"{a['product_name']} ↔ {b['product_name']}"
        url_a = a.get('source_url', '')
        url_b = b.get('source_url', '')

        alternating = (i % 2 == 0)
        style_data_row(ws, row, len(headers), alternating=alternating)

        ws.cell(row=row, column=1, value=product_label).font = DATA_FONT
        ws.cell(row=row, column=2, value=a['source_store']).font = DATA_FONT
        ws.cell(row=row, column=3, value=a['source_channel']).font = DATA_FONT

        price1_cell = ws.cell(row=row, column=4, value=price_a)
        price1_cell.number_format = MONEY_FORMAT
        price1_cell.fill = GREEN_FILL
        price1_cell.font = GREEN_FONT

        # URL 1 (col 5)
        url1_cell = ws.cell(row=row, column=5, value=url_a)
        url1_cell.font = DATA_FONT
        if url_a:
            url1_cell.hyperlink = url_a
            url1_cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')

        ws.cell(row=row, column=6, value=b['source_store']).font = DATA_FONT
        ws.cell(row=row, column=7, value=b['source_channel']).font = DATA_FONT

        price2_cell = ws.cell(row=row, column=8, value=price_b)
        price2_cell.number_format = MONEY_FORMAT
        price2_cell.font = DATA_FONT

        # URL 2 (col 9)
        url2_cell = ws.cell(row=row, column=9, value=url_b)
        url2_cell.font = DATA_FONT
        if url_b:
            url2_cell.hyperlink = url_b
            url2_cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')

        diff_cell = ws.cell(row=row, column=10, value=diff)
        diff_cell.number_format = MONEY_FORMAT
        diff_cell.font = DATA_FONT

        ws.cell(row=row, column=11, value=cheaper_at).font = DATA_FONT

    # Footer note
    footer_row = len(curated) + 6
    ws.merge_cells(f'A{footer_row}:K{footer_row}')
    ws.cell(row=footer_row, column=1,
            value='⚠ Product names vary significantly across stores. '
                  'Kaufland Glovo returned limited data (store closed at scrape time). '
                  'Verify matches manually before drawing price conclusions.')
    ws.cell(row=footer_row, column=1).font = NOTE_FONT


# ── Sheet 3: Summary ──────────────────────────────────────────────────────────
def build_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = '00B050'

    col_widths = [45, 20, 14, 14, 12, 14, 55]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    extraction_date = max((d.get('extraction_date', '') for d in data), default=date.today().isoformat())

    # Title block
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Bulgarian Grocery Promo Prices — Summary'
    ws['A1'].font = TITLE_FONT

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Extraction: {extraction_date} | Total Items: {len(data)}'
    ws['A2'].font = SUBTITLE_FONT

    # ── Section B: Coverage ────────────────────────────────────────────────────
    ws['A4'] = 'Coverage by Store & Channel'
    ws['A4'].font = SUBTITLE_FONT

    cov_headers = ['Store', 'Channel', 'Items', 'Avg Promo Price', 'Min Price', 'Max Price']
    for col, h in enumerate(cov_headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header(ws, 5, len(cov_headers))

    by_combo = defaultdict(list)
    for d in data:
        p = d.get('promo_price')
        if p:
            by_combo[(d['source_store'], d['source_channel'])].append(p)

    # Also count items without promo price
    combo_counts = Counter((d['source_store'], d['source_channel']) for d in data)

    cov_rows = sorted(combo_counts.keys(), key=lambda k: -combo_counts[k])
    for i, key in enumerate(cov_rows):
        row = i + 6
        prices = by_combo[key]
        alternating = (i % 2 == 0)
        style_data_row(ws, row, len(cov_headers), alternating=alternating)
        ws.cell(row=row, column=1, value=key[0]).font = DATA_FONT
        ws.cell(row=row, column=2, value=key[1]).font = DATA_FONT
        ws.cell(row=row, column=3, value=combo_counts[key]).font = DATA_FONT
        if prices:
            avg_cell = ws.cell(row=row, column=4, value=statistics.mean(prices))
            min_cell = ws.cell(row=row, column=5, value=min(prices))
            max_cell = ws.cell(row=row, column=6, value=max(prices))
            for c in [avg_cell, min_cell, max_cell]:
                c.number_format = MONEY_FORMAT
                c.font = DATA_FONT

    # ── Section C: Missing/failed sites ───────────────────────────────────────
    miss_start = len(cov_rows) + 8
    ws.cell(row=miss_start, column=1, value='Sites With No Data Returned').font = SUBTITLE_FONT
    ws.cell(row=miss_start, column=1).fill = RED_FILL
    ws.cell(row=miss_start, column=1).font = Font(name='Arial', bold=True, size=11, color='9C0006')

    miss_headers = ['Store', 'Channel', 'URL', 'Likely Reason']
    for col, h in enumerate(miss_headers, 1):
        ws.cell(row=miss_start + 1, column=col, value=h)
    style_header(ws, miss_start + 1, len(miss_headers))

    expected = [
        ("Kaufland", "Direct",
         "https://www.kaufland.bg/aktualni-predlozheniya/oferti.html",
         "Scraped successfully"),
        ("Billa", "Direct",
         "https://www.billa.bg/promocii/sedmichna-broshura",
         "Custom parser (ssbbilla.site mirror)"),
        ("Fantastico", "Direct",
         "https://www.fantastico.bg/special-offers",
         "PDF brochure — parsed with pdfplumber"),
        ("Kaufland", "Glovo",
         "https://glovoapp.com/bg/bg/sofia/stores/kaufland-sof?content=promotsii-pr",
         "Store closed during scrape window (opens 10:00 EET) — retry needed"),
        ("Billa", "Glovo",
         "https://glovoapp.com/bg/bg/sofia/stores/billa-sof1?content=promotsii-pr",
         "Scraped successfully"),
        ("Coca-Cola Real Magic", "Glovo",
         "https://glovoapp.com/bg/bg/sofia/stores/coca-cola-real-magic-sof",
         "Scraped successfully"),
        ("Gladen.bg / Hit Max", "Direct",
         "https://gladen.bg/promotions",
         "Scraped successfully (SSR HTML, all 42 pages)"),
    ]

    actual_urls = set(d.get('source_url', '') for d in data)

    miss_data_row = miss_start + 2
    for store, channel, url, reason in expected:
        found = any(url in u for u in actual_urls)
        status = reason if found else f"⚠ NO DATA — {reason}"
        fill = None if found else YELLOW_FILL
        for col, val in enumerate([store, channel, url, status], 1):
            c = ws.cell(row=miss_data_row, column=col, value=val)
            c.font = DATA_FONT
            if fill and col == 4:
                c.fill = fill
        miss_data_row += 1

    # ── Section D: Top 15 discounts ───────────────────────────────────────────
    top_start = miss_data_row + 2
    ws.cell(row=top_start, column=1, value='Top 15 Biggest Discounts').font = SUBTITLE_FONT

    top_headers = ['Product', 'Store', 'Regular Price', 'Promo Price', 'Discount %', 'You Save', 'URL']
    top_col_widths = [45, 20, 14, 14, 12, 14, 55]
    for col, (h, w) in enumerate(zip(top_headers, top_col_widths), 1):
        ws.cell(row=top_start + 1, column=col, value=h)
        ws.column_dimensions[ws.cell(row=top_start + 1, column=col).column_letter].width = w
    style_header(ws, top_start + 1, len(top_headers))

    discounts = []
    for d in data:
        reg   = d.get('regular_price')
        promo = d.get('promo_price')
        if reg and promo and reg > 0:
            discounts.append((1 - promo / reg, d))
    discounts.sort(reverse=True, key=lambda x: x[0])

    for i, (pct, d) in enumerate(discounts[:15]):
        row = top_start + 2 + i
        reg   = d['regular_price']
        promo = d['promo_price']
        url   = d.get('source_url', '')
        alternating = (i % 2 == 0)
        style_data_row(ws, row, len(top_headers), alternating=alternating)

        ws.cell(row=row, column=1, value=d['product_name']).font = DATA_FONT
        ws.cell(row=row, column=2, value=d['source_store']).font = DATA_FONT
        for col, val in [(3, reg), (4, promo), (6, reg - promo)]:
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = MONEY_FORMAT
            c.font = DATA_FONT
        disc_cell = ws.cell(row=row, column=5, value=pct)
        disc_cell.number_format = PCT_FORMAT
        disc_cell.fill = GREEN_FILL
        disc_cell.font = GREEN_FONT

        url_cell = ws.cell(row=row, column=7, value=url)
        if url:
            url_cell.hyperlink = url
            url_cell.font = Font(name='Arial', size=10, color='0563C1', underline='single')
        else:
            url_cell.font = DATA_FONT

    print(f"Sheet 3 'Summary': written.")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading {INPUT_PATH}...")
    with open(INPUT_PATH, encoding='utf-8') as f:
        data = json.load(f)

    run_quality_checks(data)

    # Compute extraction date for filename
    extraction_date = max(
        (d.get('extraction_date', '') for d in data),
        default=date.today().isoformat()
    )
    output_path = BASE_DIR / f"bg_promo_prices_{extraction_date}.xlsx"

    print("\nBuilding XLSX...")
    wb = Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    build_all_items_sheet(wb, data)
    build_cross_store_sheet(wb, data)
    build_summary_sheet(wb, data)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024:.1f} KB")


if __name__ == '__main__':
    main()
