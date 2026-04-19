#!/usr/bin/env python3
"""
retrofit_urls.py — Inject source_url hyperlinks into older Excel files that
were generated before URL columns were added to generate_cheapest_xlsx.py.

Usage:
    python retrofit_urls.py bg_cheapest_by_category_2026-03-29.xlsx

The script:
  1. Builds a URL lookup from bulgarian_promo_prices_merged.json keyed on
     (product_name.lower(), source_store.lower()) — tolerates week-over-week
     product name variations as long as store matches.
  2. For each sheet/row that has a product-name + store pair but no URL,
     appends a URL column (or fills an existing empty one) with a hyperlink.
  3. Saves the result as <original>_with_urls.xlsx.
"""

import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

BASE        = Path(__file__).parent
MASTER_PATH = BASE / "bulgarian_promo_prices_merged.json"
URL_FONT    = Font(name="Arial", size=10, color="0563C1", underline="single")

# ── Build URL lookup ──────────────────────────────────────────────────────────

def build_lookup(master_path: Path) -> dict:
    """
    Returns dict keyed by (product_name_lower, store_lower) → source_url.
    Also adds a fallback keyed by (product_name_lower,) for cases where
    the store column isn't easily parseable.
    """
    with open(master_path, encoding="utf-8") as f:
        records = json.load(f)

    lookup = {}
    for r in records:
        name  = (r.get("product_name") or "").strip().lower()
        store = (r.get("source_store") or "").strip().lower()
        url   = r.get("source_url") or ""
        if name and url:
            lookup[(name, store)] = url
            # fallback without store
            if (name,) not in lookup:
                lookup[(name,)] = url
    return lookup


def find_url(lookup, product_name, store=""):
    name  = (product_name or "").strip().lower()
    store = (store or "").strip().lower()
    return (lookup.get((name, store))
            or lookup.get((name,))
            or "")


# ── Sheet-specific retrofit logic ─────────────────────────────────────────────

def _col_index(headers: list, *candidates) -> int | None:
    """Return 1-based column index of the first matching header name."""
    for h in headers:
        for candidate in candidates:
            if h and candidate.lower() in str(h).lower():
                return headers.index(h) + 1
    return None


def retrofit_sheet(ws, lookup, product_col, store_col, url_col_hint=None,
                   data_start_row=2):
    """
    For each data row in ws, look up the URL and write it.
    product_col / store_col / url_col_hint are 1-based column indices.
    If url_col_hint is None, appends a new column after the last used column.
    Returns count of URLs written.
    """
    max_col = ws.max_column or 1
    url_col = url_col_hint or (max_col + 1)

    # Write "URL" header just before the data rows
    header_cell = ws.cell(row=data_start_row - 1, column=url_col)
    if not (header_cell.value and "url" in str(header_cell.value).lower()):
        header_cell.value = "URL"

    written = 0
    for row in ws.iter_rows(min_row=data_start_row):
        try:
            name_cell  = row[product_col - 1]
            store_cell = row[store_col - 1] if store_col else None
            name       = str(name_cell.value or "").strip()
            store      = str(store_cell.value or "").strip() if store_cell else ""
        except IndexError:
            continue

        if not name or len(name) < 4:
            continue

        url = find_url(lookup, name, store)
        if not url:
            continue

        url_cell = ws.cell(row=name_cell.row, column=url_col)
        if url_cell.value:
            continue  # already populated — don't overwrite
        url_cell.value     = url
        url_cell.hyperlink = url
        url_cell.font      = URL_FONT
        written += 1

    return written


# ── Main ──────────────────────────────────────────────────────────────────────

def retrofit_file(xlsx_path: Path, lookup: dict) -> Path:
    print(f"Opening: {xlsx_path.name}")
    wb = load_workbook(str(xlsx_path))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect all header values from first 5 rows to find column positions
        headers = []
        header_row_idx = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6,
                                                    values_only=False), 1):
            vals = [c.value for c in row]
            # A header row has multiple non-null text values
            text_count = sum(1 for v in vals if v and isinstance(v, str) and len(v) > 1)
            if text_count >= 3:
                headers       = [c.value for c in row]
                header_row_idx = row_idx
                break

        if not headers:
            print(f"  Sheet '{sheet_name}': no header row found, skipping.")
            continue

        # Locate product-name and store columns
        product_col = _col_index(headers, "Продукт", "Най-евтин", "product")
        store_col   = _col_index(headers, "Магазин", "store")
        url_col     = _col_index(headers, "URL", "url", "линк")

        if not product_col:
            print(f"  Sheet '{sheet_name}': no product column found, skipping.")
            continue

        count = retrofit_sheet(ws, lookup, product_col, store_col, url_col,
                               data_start_row=(header_row_idx or 1) + 1)
        print(f"  Sheet '{sheet_name}': {count} URLs written "
              f"(product col={product_col}, store col={store_col}, url col={url_col or 'new'})")

    out_path = xlsx_path.with_stem(xlsx_path.stem + "_with_urls")
    wb.save(str(out_path))
    print(f"\nSaved: {out_path.name}")
    return out_path


def main():
    targets = sys.argv[1:]
    if not targets:
        # Default to the March 29 file
        targets = [str(BASE / "bg_cheapest_by_category_2026-03-29.xlsx")]

    lookup = build_lookup(MASTER_PATH)
    print(f"URL lookup: {len(lookup)} entries from master JSON\n")

    for t in targets:
        p = Path(t)
        if not p.is_absolute():
            p = BASE / p
        if not p.exists():
            print(f"Not found: {p}")
            continue
        retrofit_file(p, lookup)


if __name__ == "__main__":
    main()
