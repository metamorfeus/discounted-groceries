#!/usr/bin/env python3
"""
Billa Bulgaria — Weekly PDF Brochure Pipeline
==============================================
Source: https://www.billa.bg/promocii/sedmichna-broshura
        (Publitas-hosted flipbook; image-only PDF — requires Azure OCR)

Pipeline steps:
  1. Scrape billa.bg → find the current week's Publitas viewer URL
  2. Extract the direct PDF download URL from the viewer page source
       Playwright fallback: if direct URL not found, launch headless browser
       and click "Изтегляне на PDF"
  3. Split PDF into BILLA_PAGES_PER_BATCH-page batches (default 2)
  4. OCR each batch via Azure Document Intelligence prebuilt-read (cached)
  5. Parse OCR text → structured product records with EUR prices
  6. If BILLA_WEEKLY_COMPARISON is enabled (config.py):
       a. Fetch current ssbbilla.site data
       b. Fuzzy-match (difflib) PDF products vs ssbbilla products
       c. Save weekly comparison report → billa_work/comparison_YYYY-MM-DD.xlsx
  7. Items from the PDF that are NOT found on ssbbilla.site are added to the
     master JSON as supplementary Billa Direct records.

Usage:
  # Full pipeline (auto-download + OCR + compare + merge)
  python billa_pdf_pipeline.py --key YOUR_AZURE_KEY

  # Use a PDF you already downloaded
  python billa_pdf_pipeline.py --key YOUR_AZURE_KEY --pdf "billa_work/brochure.pdf"

  # Reuse cached OCR output (skip download & OCR)
  python billa_pdf_pipeline.py --ocr-dir billa_work/ocr_output/

  # Dry run — parse and report, do not write to master JSON
  python billa_pdf_pipeline.py --key YOUR_AZURE_KEY --dry-run

Requirements:
  pip install requests PyPDF2 openpyxl azure-ai-documentintelligence
  # For Playwright fallback:
  pip install playwright && playwright install chromium
"""

import json
import os
import re
import sys
import time
import argparse
import logging
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ─── Load config ──────────────────────────────────────────────────────────────
try:
    import config as _cfg
    AZURE_ENDPOINT      = _cfg.AZURE_ENDPOINT
    AZURE_MODEL_ID      = _cfg.AZURE_MODEL_ID
    PAGES_PER_BATCH     = _cfg.BILLA_PAGES_PER_BATCH
    MAX_RETRIES         = _cfg.BILLA_OCR_MAX_RETRIES
    RETRY_DELAY         = _cfg.BILLA_OCR_RETRY_DELAY
    WORK_DIR            = Path(_cfg.BILLA_WORK_DIR)
    MASTER_PATH         = Path(_cfg.MASTER_JSON)
    WEEKLY_COMPARISON   = _cfg.BILLA_WEEKLY_COMPARISON
    COMPARISON_THRESHOLD = _cfg.BILLA_COMPARISON_THRESHOLD
except ImportError:
    log.warning("config.py not found — using built-in defaults")
    AZURE_ENDPOINT       = "https://invoice2024.cognitiveservices.azure.com/"
    AZURE_MODEL_ID       = "prebuilt-read"
    PAGES_PER_BATCH      = 2
    MAX_RETRIES          = 3
    RETRY_DELAY          = 5
    WORK_DIR             = Path("billa_work")
    MASTER_PATH          = Path("bulgarian_promo_prices_merged.json")
    WEEKLY_COMPARISON    = True
    COMPARISON_THRESHOLD = 0.80

# ─── Constants ────────────────────────────────────────────────────────────────
BILLA_BG_URL   = "https://www.billa.bg/promocii/sedmichna-broshura"
PUBLITAS_ACCOUNT = "billa-bulgaria"
SOURCE_STORE   = "Billa"
SOURCE_CHANNEL = "Direct"
SOURCE_URL     = BILLA_BG_URL
EXTRACTION_DATE = date.today().isoformat()
EUR_TO_BGN     = 1.95583

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "bg-BG,bg;q=0.9,en;q=0.8",
}

# ─── OCR regex patterns — tuned for Billa's OCR noise ────────────────────────
#
# Key OCR issues in Billa's image PDF:
#   • '€' is frequently read as '6', 'Є', 'E', or 'б'
#   • BGN trailing period is often absent:  "6.83 ЛВ" not "6.83 ЛВ."
#   • BGN has only 1 decimal digit:         "19.9 ЛВ"  (= 19.90)
#   • Hyphen used as decimal in BGN:        "6-83 ЛВ"  → "6.83 ЛВ"
#   • Space inside BGN number:              "6 43 ЛВ"  → "6.43 ЛВ"
#   • 'лв' misread as 'JB' or 'MB'
#   • Bullet used as decimal:               "10•93 ЛВ" → "10.93 ЛВ"
#   • Old EUR price often garbled;
#     discount % ("-33%") is read correctly → used to back-calculate regular price
#
# Strategy: BGN is the reliable anchor. promo_eur = round(bgn / 1.95583, 2).
# Regular_eur derived from discount % when available; clean EUR pair as fallback.

# BGN price — handles missing period, 1 decimal, comma separator, JB/MB misread
_BGN_RE = re.compile(
    r"(\d{1,3}[.,]\d{1,2})\s*(?:ЛВ|лв|JB|MB)\.?(?!\d)",
    re.IGNORECASE,
)

# EUR price — handles correct € AND common OCR misreads (6, Є, Е)
# Requires exactly 2 decimal places to reduce false positives from digit '6'
_EUR_RE = re.compile(
    r"(\d{1,3}[.,]\d{2})\s*[€ЄЕ]|"      # correct € or Є/Е
    r"(\d{1,3}[.,]\d{2})\s+6(?!\d)",      # '6' as '€', not followed by digit
    re.IGNORECASE,
)

# Discount percentage: "-33%" or "- 47%"
_DISC_RE = re.compile(r"-\s*(\d{1,2})\s*%")

# BILLA Card regular price: "цена без BILLA Card 8,18 €"
_CARD_PRICE_RE = re.compile(
    r"цена без BILLA Card\s+(\d{1,3}[.,]\d{2})\s*[€ЄЕ6Є]",
    re.IGNORECASE,
)

_UNIT_RE   = re.compile(r"цена\s+за\s+(бр|кг|оп|л|к-кт)\.?", re.IGNORECASE)
_WEIGHT_RE = re.compile(r"(\d+(?:[,.]\d+)?)\s*(кг|г|мл|л)\b", re.IGNORECASE)

_SKIP_RE = re.compile(
    r"^(?:"
    r"стр\.\s*\d|billa\.bg|www\.|ОФЕРТА ЗА|Продуктите се продават|"
    r"\d{1,3}$|℮|Всички\s+.*(?:с\s+марка|видове)|минимум\s+-?\d+|"
    r"не носи отговорност|Алергени:|catering@|тел\.\s*\+|"
    r"\*+|Период на кампанията|предложението е валидно|акцията е валидна|"
    r"^•|^·|^\)|^\[|^-{2,}|^/|Валидност|BILLA\s+Card|"
    r"[EeЄЕ€]?\s*СУПЕР ЦЕНА|СУПЕР ЦЕНА|СЕГА В BILLA|"
    r"Супер цена|Сега в Billa|Само с Billa|Мултипак|Ново в Billa|"
    r"Color Week|Най-добра цена|"
    r"Кликни тук|BILLA ready|От топлата витрина|От BILLA пекарна|"
    r"От деликатесната витрина|Вдохнови се|Извинете ни|"
    r"BILLA ВИНАГИ|ЧЕСТИТА|до изчерпване|"
    r"Посочените цени|по [Вв]алутен курс|[Цц]ените са [Вв]алидни|"
    r"[Ии]зображенията са|[Аа]ртикулът не е наличен|"
    r"1 изпиране|цена без BILLA|Цена за \d|Цена за 1|"
    r"SUPER PROMO|Card$|^Card\b|"
    r"Офертата е|Офертата е [Вв]алидна|до 08\.|до 05\.|"
    r"ПРАЗНИЧНИ|TPAAVUMOKHA|МАТУРАЛЕН|"
    r"цена при оборот|цена без натрупан оборот|"
    r"направени в периода|СПЕСТЯВАЙ|месечни покупки|"
    r"Кликни върху|при следваща покупка|Ваучер за еднократна|"
    r"само с твоята Card|при достигане на оборот|отстъпка за мин\.|"
    r"Пазарувай В периода|е минимум \d|ДОВЕРИ СЕ НА|"
    r"\d+°\s*[СC]\s*\||Отключи отстъпки|без собиране на точки|"
    r"В периода от \d{2}\.|Виж продуктовите ваучери|"
    r"Свали BILLA App|ЕТО И ПРОДУКТИТЕ|Избери Domenico|"
    r"Пазарувай за|да почистиш всичко|"
    r"BILLA ДОВЕРИ СЕ НА|OT 200%|Чистота за пример|"
    r"^,\s*е минимум|^,\s*има право|Ha Bcuyku|Ha Bс[иi]чки"
    r")",
    re.IGNORECASE,
)
_CYRILLIC_RE = re.compile(r"[\u0400-\u04FF]")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Find current week's Publitas viewer URL
# ══════════════════════════════════════════════════════════════════════════════

def find_publitas_viewer_url(session):
    """
    Scrape billa.bg/promocii/sedmichna-broshura and extract the
    Publitas viewer URL for the current week's brochure.

    Returns (viewer_url, slug) or (None, None) if not found.
    """
    log.info(f"Scraping {BILLA_BG_URL} for current brochure link...")
    try:
        resp = session.get(BILLA_BG_URL, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        log.error(f"Could not fetch billa.bg: {e}")
        return None, None

    html = resp.text

    # Publitas viewer URLs appear in iframes, data-src attributes, or JS
    patterns = [
        rf"https://view\.publitas\.com/{PUBLITAS_ACCOUNT}/([^\"'\s/]+)",
        rf"//view\.publitas\.com/{PUBLITAS_ACCOUNT}/([^\"'\s/]+)",
    ]
    for pat in patterns:
        m = re.search(pat, html)
        if m:
            slug = m.group(1).rstrip("/")
            # Strip any page-path suffix: "/page/1"
            slug = re.sub(r"/page/\d+$", "", slug)
            viewer_url = f"https://view.publitas.com/{PUBLITAS_ACCOUNT}/{slug}/page/1"
            log.info(f"Found Publitas slug: {slug}")
            return viewer_url, slug

    log.error(
        "Could not find Publitas link on billa.bg.\n"
        "The page structure may have changed. "
        "Try passing --pdf with a manually downloaded file."
    )
    return None, None


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Download PDF
# ══════════════════════════════════════════════════════════════════════════════

def _extract_pdf_url_from_viewer_html(html):
    """
    Search the Publitas viewer page source for the direct PDF download URL.
    Pattern: https://view.publitas.com/ACCOUNT_ID/PUB_ID/pdfs/UUID.pdf
    """
    # Unescaped form
    m = re.search(
        r"https://view\.publitas\.com/\d+/\d+/pdfs/[a-f0-9\-]+\.pdf"
        r"(?:\?[^\"'<\s]*)?",
        html, re.IGNORECASE,
    )
    if m:
        return m.group(0)

    # JSON-escaped form  (\/ instead of /)
    m2 = re.search(
        r"https:\\/\\/view\\.publitas\\.com\\/\d+\\/\d+\\/pdfs\\/[a-f0-9\\-]+\\.pdf",
        html, re.IGNORECASE,
    )
    if m2:
        return m2.group(0).replace("\\/", "/")

    return None


def download_pdf_direct(session, viewer_url, output_path):
    """
    Fetch the Publitas viewer page, extract the embedded PDF URL, download it.
    Returns output_path on success, None on failure.
    """
    log.info(f"Fetching viewer page to extract PDF URL: {viewer_url}")
    try:
        resp = session.get(viewer_url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        log.warning(f"Could not fetch viewer page: {e}")
        return None

    pdf_url = _extract_pdf_url_from_viewer_html(resp.text)
    if not pdf_url:
        log.warning("PDF URL not found in viewer page source.")
        return None

    log.info(f"Found PDF URL: {pdf_url[:80]}...")
    try:
        pdf_resp = session.get(pdf_url, timeout=120, stream=True)
        pdf_resp.raise_for_status()
        with open(output_path, "wb") as f:
            for chunk in pdf_resp.iter_content(chunk_size=65536):
                f.write(chunk)
        size_mb = os.path.getsize(output_path) / (1024 * 1024)
        # Verify it's actually a PDF
        with open(output_path, "rb") as f:
            if f.read(4) != b"%PDF":
                log.error("Downloaded file is not a valid PDF — deleting")
                os.remove(output_path)
                return None
        log.info(f"PDF downloaded: {output_path} ({size_mb:.1f} MB)")
        return output_path
    except Exception as e:
        log.warning(f"Direct PDF download failed: {e}")
        if os.path.exists(output_path):
            os.remove(output_path)
        return None


def download_pdf_playwright(viewer_url, output_path):
    """
    Fallback: launch a headless Chromium browser via Playwright,
    navigate to the Publitas viewer, and click 'Изтегляне на PDF'.
    Returns output_path on success, None on failure.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.error(
            "Playwright not installed. Install with:\n"
            "  pip install playwright && playwright install chromium"
        )
        return None

    log.info("Falling back to Playwright browser automation...")
    download_dir = str(WORK_DIR.resolve())

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()

            log.info(f"Navigating to: {viewer_url}")
            page.goto(viewer_url, timeout=60000)
            page.wait_for_load_state("networkidle", timeout=30000)

            # Try to click the download button (Bulgarian label)
            selectors = [
                'text="Изтегляне на PDF"',
                'text="Изтегли PDF"',
                '[aria-label*="PDF"]',
                '[title*="PDF"]',
                'button:has-text("PDF")',
                'a:has-text("PDF")',
            ]
            clicked = False
            for sel in selectors:
                try:
                    if page.locator(sel).count() > 0:
                        with page.expect_download(timeout=60000) as dl_info:
                            page.locator(sel).first.click()
                        dl = dl_info.value
                        dl.save_as(output_path)
                        clicked = True
                        log.info(f"PDF downloaded via Playwright: {output_path}")
                        break
                except Exception:
                    continue

            browser.close()

            if not clicked:
                log.error(
                    "Could not find download button in Publitas viewer.\n"
                    "Please download the PDF manually and pass it with --pdf"
                )
                return None

        # Verify
        with open(output_path, "rb") as f:
            if f.read(4) != b"%PDF":
                log.error("Playwright download is not a valid PDF")
                os.remove(output_path)
                return None
        return output_path

    except Exception as e:
        log.error(f"Playwright error: {e}")
        return None


def download_pdf(viewer_url, output_path, session):
    """
    Attempt direct download, fall back to Playwright.
    Returns output_path on success, None on failure.
    """
    result = download_pdf_direct(session, viewer_url, output_path)
    if result:
        return result
    log.info("Direct download failed — trying Playwright fallback")
    return download_pdf_playwright(viewer_url, output_path)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Split PDF into batches
# ══════════════════════════════════════════════════════════════════════════════

def split_pdf(pdf_path, batch_dir, pages_per_batch=PAGES_PER_BATCH):
    """
    Split the PDF into small batches for Azure DI.
    Returns list of batch file paths.
    """
    from PyPDF2 import PdfReader, PdfWriter

    batch_dir = Path(batch_dir)
    batch_dir.mkdir(parents=True, exist_ok=True)

    reader = PdfReader(str(pdf_path))
    total = len(reader.pages)
    log.info(f"PDF has {total} pages — splitting into {pages_per_batch}-page batches")

    batches = []
    for start in range(0, total, pages_per_batch):
        end = min(start + pages_per_batch, total)
        num = start // pages_per_batch + 1
        out = batch_dir / f"batch_{num:03d}_pages_{start+1}-{end}.pdf"

        writer = PdfWriter()
        for i in range(start, end):
            writer.add_page(reader.pages[i])
        with open(out, "wb") as f:
            writer.write(f)

        batches.append(str(out))

    log.info(f"Created {len(batches)} batches in {batch_dir}")
    return batches


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — OCR via Azure Document Intelligence (with caching)
# ══════════════════════════════════════════════════════════════════════════════

def ocr_batch(pdf_path, endpoint, key):
    """Send one PDF batch to Azure DI and return the result dict."""
    from azure.core.credentials import AzureKeyCredential
    from azure.ai.documentintelligence import DocumentIntelligenceClient

    client = DocumentIntelligenceClient(endpoint, AzureKeyCredential(key))
    with open(pdf_path, "rb") as f:
        data = f.read()

    log.info(f"  OCR: {Path(pdf_path).name} ({len(data)/1024:.0f} KB)")
    try:
        poller = client.begin_analyze_document(
            AZURE_MODEL_ID, body=data, content_type="application/pdf"
        )
    except TypeError:
        poller = client.begin_analyze_document(
            AZURE_MODEL_ID,
            analyze_request=data,
            content_type="application/pdf",
        )
    result = poller.result()

    pages_data = []
    for pg in result.pages:
        pages_data.append({
            "page_number": pg.page_number,
            "text": "\n".join(ln.content for ln in pg.lines),
            "lines": [ln.content for ln in pg.lines],
            "words": [{"content": w.content, "confidence": w.confidence} for w in pg.words],
        })

    return {"pages": pages_data, "full_text": result.content or "", "source_file": pdf_path}


def ocr_all_batches(batch_files, ocr_dir, endpoint, key):
    """
    OCR all batches, caching results as JSON in ocr_dir.
    Already-processed batches are loaded from cache.
    Returns list of OCR result dicts.
    """
    ocr_dir = Path(ocr_dir)
    ocr_dir.mkdir(parents=True, exist_ok=True)
    results = []

    for i, batch_path in enumerate(batch_files):
        name = Path(batch_path).stem
        cache = ocr_dir / f"{name}_ocr.json"

        if cache.exists():
            log.info(f"  Batch {i+1}/{len(batch_files)}: {name} — loading from cache")
            with open(cache, encoding="utf-8") as f:
                results.append(json.load(f))
            continue

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                res = ocr_batch(batch_path, endpoint, key)
                with open(cache, "w", encoding="utf-8") as f:
                    json.dump(res, f, ensure_ascii=False, indent=2)
                log.info(
                    f"  Batch {i+1}/{len(batch_files)}: {name} "
                    f"— {len(res['pages'])} pages OCR'd"
                )
                results.append(res)
                break
            except Exception as e:
                log.warning(f"  Batch {i+1} attempt {attempt}/{MAX_RETRIES} failed: {e}")
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                else:
                    log.error(f"  Batch {i+1} FAILED after {MAX_RETRIES} attempts — skipping")

        if i < len(batch_files) - 1:
            time.sleep(1)  # avoid rate-limit bursts

    return results


def load_ocr_from_directory(ocr_dir):
    """Load all cached OCR JSON files from a directory."""
    ocr_dir = Path(ocr_dir)
    files = sorted(ocr_dir.glob("*_ocr.json"))
    if not files:
        log.error(f"No OCR JSON files found in {ocr_dir}")
        return []
    results = []
    for f in files:
        with open(f, encoding="utf-8") as fh:
            results.append(json.load(fh))
    log.info(f"Loaded {len(results)} cached OCR batches from {ocr_dir}")
    return results


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Parse OCR text → product records
# ══════════════════════════════════════════════════════════════════════════════

def _is_name_line(line):
    line = line.strip()
    if not line or len(line) < 4:
        return False
    if _SKIP_RE.match(line):
        return False
    if re.match(r"^[\d.]+\s*[€%]", line):
        return False
    if re.match(r"^-?\d+%", line):
        return False
    if _BGN_RE.match(line):
        return False
    if re.match(r"^цена\s+за", line, re.IGNORECASE):
        return False
    if re.match(r"^\d{1,3}$", line):
        return False
    if re.match(r"^\d+(?:[,.]\d+)?\s*(?:г|кг|мл|л|бр)\.?$", line, re.IGNORECASE):
        return False
    # Short all-caps Latin fragments (logo text, brand stamps)
    if re.match(r"^[A-Z0-9\s.\-!?]+$", line) and len(line) < 12:
        return False
    # All-caps Cyrillic short words are section headers/promo labels, not product names
    # e.g. "КЛАСИК", "ПРАЗНИЧНИ", "МУЛТИ ПАК", "ОФЕРТИ"
    if re.match(r"^[А-ЯЁ\s\-]+$", line) and len(line) < 20:
        return False
    return True


def _name_score(line):
    first = next((c for c in line if c.isalpha()), None)
    return 2 if (first and "\u0400" <= first <= "\u04FF") else 1


def _clean_name(name):
    """Strip known noise suffixes from a product name."""
    name = re.sub(r"\s*цена за.*$",                        "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*Цена за \d+\s*(?:бр|р)\.?.*$",    "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*(?:цена без|6e3)\s+BILLA\b.*$",     "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*цена при оборот.*$",                "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*произход\s+\S+.*$",                "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*BILLA ready\s*$",                  "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*От топлата витрина\s*$",           "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*От BILLA пекарна\s*$",             "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*От деликатесната витрина\s*$",     "", name, flags=re.IGNORECASE)
    # Strip trailing OCR noise: "ДОЖДИК -" → "ДОЖДИК", "Product -33%" → "Product"
    name = re.sub(r"\s+[-–]\s*$",                         "", name)
    name = re.sub(r"\s+-\d+%\s*$",                        "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def _is_valid_name(name):
    """Return True only if name has 4+ consecutive Cyrillic chars (filters garbage OCR)."""
    return bool(re.search(r"[\u0400-\u04FF]{4,}", name))


def _extract_name(region):
    """Extract product name from text that appears BEFORE the price."""
    lines = [l.strip() for l in region.split("\n")][-6:]
    scored = [(l, _name_score(l)) for l in lines if _is_name_line(l)]
    if not scored:
        return None
    cyrillic = [l for l, s in scored if s == 2]
    pool = cyrillic if cyrillic else [l for l, _ in scored]
    best = max(pool, key=len)
    best = _clean_name(best)
    if len(best) < 8 or not _is_valid_name(best):
        return None
    return best


def _extract_name_forward(region):
    """
    Extract product name from text that appears AFTER the price.
    Billa OCR often reads the price block first, then the product name below it.

    Takes the FIRST valid name line (closest to the price), not the longest —
    the first meaningful Cyrillic line is most likely to belong to this product,
    while later lines belong to the next product's pre-price text block.
    """
    lines = [l.strip() for l in region.split("\n")][:6]
    for line in lines:
        if not _is_name_line(line):
            continue
        cleaned = _clean_name(line)
        if len(cleaned) >= 10 and _is_valid_name(cleaned):
            return cleaned
    return None


def _extract_unit(text):
    m = _UNIT_RE.search(text)
    if m:
        return {"кг": "кг", "бр": "бр", "оп": "опаковка",
                "л": "л", "к-кт": "к-кт"}.get(m.group(1).lower())
    w = _WEIGHT_RE.search(text)
    if w:
        return f"{w.group(1)} {w.group(2)}"
    return None


def _auto_categorize(name):
    n = name.lower()
    if "маслиново масло" in n or "маслини" in n or "олио" in n:
        return "Масла и подправки"
    cats = {
        "Месо":    ["месо", "пиле", "пилеш", "свинск", "телеш", "агнеш", "кебапч",
                    "кюфте", "наденица", "суджук", "шунка", "салам", "бут", "плешка",
                    "вешалица", "джолан", "крила", "бекон", "кренвирш"],
        "Риба":    ["риба", "рибен", "сьомга", "хек", "скумрия", "скарид"],
        "Млечни продукти": ["сирене", "масло", "мляко", "кисело", "кашкавал",
                            "йогурт", "извара", "сметана", "крема"],
        "Плодове и зеленчуци": ["грозде", "банан", "ябълк", "домат", "краставиц",
                                 "картоф", "лук", "салата", "авокадо", "портокал",
                                 "лимон", "череш", "ягод", "диня", "пъпеш", "чушк"],
        "Напитки": ["бира", "вино", "сок", "вода", "напитка", "кафе", "чай",
                    "кола", "пепси", "ракия", "уиски", "водка"],
        "Хляб и тестени": ["хляб", "сомун", "кифл", "баница", "козунак",
                            "питка", "бисквит", "вафл", "тост", "кроасан"],
        "Домакинство": ["тоалетна хартия", "кухненска ролка", "пране", "почиств",
                        "препарат", "гел за", "прах за", "омекотител", "кърпа"],
        "Консерви": ["пастет", "лютеница", "стерилизиран", "компот", "консерв"],
    }
    for cat, kws in cats.items():
        if any(kw in n for kw in kws):
            return cat
    return None


def _parse_bgn(s):
    """Convert BGN string (may use comma or period as decimal) to float."""
    return float(s.replace(",", "."))


def parse_text_stream(text, promo_period):
    """
    Parse a continuous OCR text stream into Billa product records.

    Billa-specific approach (differs from Fantastico):
      • BGN price is the reliable anchor — '€' is frequently OCR'd as '6'
      • promo_eur is ALWAYS derived from BGN: round(bgn / 1.95583, 2)
      • regular_eur priority:
          1. Explicit 'цена без BILLA Card X,XX €' in the window
          2. Discount % back-calculation: promo_eur / (1 - disc/100)
          3. Clean EUR value clearly above promo_eur in the window
      • Product name: last meaningful Cyrillic line(s) before the price block

    All prices stored in EUR.
    """
    products = []
    seen = set()

    # ── Pre-process OCR artefacts ─────────────────────────────────────────────

    # Remove "цена без BILLA Card X,XX € / Y,YY ЛВ" — the BGN here is the
    # REGULAR price, not the promo price. Strip it so it doesn't become a BGN anchor.
    text = re.sub(
        r"цена без BILLA Card\s+\d+[.,]\d+\s*[€ЄЕ6]\s*/\s*\d+[.,]\d+\s*(?:лв|ЛВ)\.?",
        "цена без BILLA Card [removed]",
        text, flags=re.IGNORECASE,
    )

    # Fix hyphen as decimal in BGN — allow 1 OR 2 digits after hyphen:
    # "6-83 ЛВ" → "6.83 ЛВ",  "39-8 ЛВ" → "39.8 ЛВ"
    text = re.sub(
        r"\b(\d{1,3})-(\d{1,2})\s*(?:ЛВ|лв|JB|MB)",
        lambda m: f"{m.group(1)}.{m.group(2)} ЛВ",
        text, flags=re.IGNORECASE,
    )
    # Fix bullet/middle-dot as decimal: "10•93 ЛВ" → "10.93 ЛВ"
    text = re.sub(
        r"\b(\d+)[•·](\d{1,2})\s*(?:ЛВ|лв|JB|MB)",
        lambda m: f"{m.group(1)}.{m.group(2)} ЛВ",
        text, flags=re.IGNORECASE,
    )
    # Fix space inside BGN number before ЛВ: "6 43 ЛВ" → "6.43 ЛВ"
    text = re.sub(
        r"\b(\d)\s+(\d{2})\s+(?:ЛВ|лв|JB|MB)",
        lambda m: f"{m.group(1)}.{m.group(2)} ЛВ",
        text, flags=re.IGNORECASE,
    )
    # Normalise JB / MB → ЛВ (OCR misread of 'лв')
    text = re.sub(r"\b(JB|MB)\.?", "ЛВ", text)

    def _emit(name, promo_eur, regular_eur, unit):
        if not name or promo_eur < 0.60 or promo_eur > 260:
            return
        key = (name[:40].lower(), promo_eur)
        if key in seen:
            return
        seen.add(key)
        products.append({
            "source_store":     SOURCE_STORE,
            "source_channel":   SOURCE_CHANNEL,
            "product_name":     name,
            "product_category": _auto_categorize(name),
            "regular_price":    regular_eur,
            "promo_price":      promo_eur,
            "unit":             unit,
            "price_per_unit":   None,
            "promo_period":     promo_period,
            "source_url":       SOURCE_URL,
            "extraction_date":  EXTRACTION_DATE,
        })

    def _regular_eur_from_window(window, promo_eur):
        """
        Try to determine the regular/old EUR price from the product's text window.
        Priority:
          1. Explicit 'цена без BILLA Card X,XX' price
          2. Discount % back-calculation
          3. Clean EUR value that is clearly higher than promo_eur
        Returns float or None.
        """
        # 1. BILLA Card explicit regular price
        card_m = _CARD_PRICE_RE.search(window)
        if card_m:
            try:
                val = float(card_m.group(1).replace(",", "."))
                if val > promo_eur:
                    return val
            except ValueError:
                pass

        # 2. Discount % → back-calculate old price
        disc_m = _DISC_RE.search(window)
        if disc_m:
            disc = int(disc_m.group(1))
            if 5 <= disc <= 70:           # realistic grocery discount range
                return round(promo_eur / (1 - disc / 100), 2)

        # 3. Look for a clean EUR value (with actual € symbol) above promo_eur
        for m in _EUR_RE.finditer(window):
            raw = m.group(1) or m.group(2)
            if not raw:
                continue
            try:
                val = float(raw.replace(",", "."))
                if val > promo_eur * 1.05:
                    return val
            except ValueError:
                pass

        return None

    # ── Pass 1: BGN-anchored ──────────────────────────────────────────────────
    bgn_matches = list(_BGN_RE.finditer(text))

    for idx, bgn_m in enumerate(bgn_matches):
        try:
            bgn_price = _parse_bgn(bgn_m.group(1))
        except ValueError:
            continue
        if bgn_price < 0.20 or bgn_price > 980:
            continue

        bgn_pos      = bgn_m.start()
        window_start = bgn_matches[idx - 1].end() if idx > 0 else max(0, bgn_pos - 600)
        window       = text[window_start:bgn_pos]

        # promo_eur derived from BGN — always reliable
        promo_eur = round(bgn_price / EUR_TO_BGN, 2)

        # regular_eur from window context
        regular_eur = _regular_eur_from_window(window, promo_eur)

        # Name boundary: end just before the first EUR-like value or unit indicator
        name_end = bgn_pos
        unit_m = _UNIT_RE.search(window)
        if unit_m:
            name_end = window_start + unit_m.start()
        else:
            first_eur = _EUR_RE.search(window)
            if first_eur:
                name_end = window_start + first_eur.start()

        unit = _extract_unit(window)

        # Try backward name (text before price) — primary strategy
        name = _extract_name(text[window_start:name_end])

        # Forward name is a fallback ONLY: used when backward extraction fails.
        # Billa OCR sometimes reads price first, then product name below.
        # We take only the FIRST valid line after the price (closest to it),
        # to avoid picking up the next product's pre-price description.
        if not name:
            next_bgn_start = (bgn_matches[idx + 1].start()
                              if idx + 1 < len(bgn_matches)
                              else bgn_m.end() + 250)
            forward_region = text[bgn_m.end(): min(bgn_m.end() + 250, next_bgn_start)]
            name = _extract_name_forward(forward_region)

        _emit(name, promo_eur, regular_eur, unit)

    return products


def parse_ocr_results(ocr_results, promo_period):
    """Parse all OCR batch results and deduplicate."""
    products = []
    seen = set()

    for batch in ocr_results:
        full_text = batch.get("full_text", "")
        if not full_text:
            full_text = "\n".join(pg.get("text", "") for pg in batch.get("pages", []))
        if not full_text.strip():
            continue
        for p in parse_text_stream(full_text, promo_period):
            key = (p["product_name"][:40].lower(), p["promo_price"])
            if key not in seen:
                seen.add(key)
                products.append(p)

    return products


def extract_promo_period(ocr_results):
    """
    Try to find the brochure validity period (e.g. '02.04 - 08.04.2026')
    in the OCR text. Falls back to today's date range string.
    """
    period_re = re.compile(
        r"(\d{2}\.\d{2}\.?\d{0,4})\s*[-–—]\s*(\d{2}\.\d{2}\.\d{4})",
        re.IGNORECASE,
    )
    for batch in ocr_results:
        text = batch.get("full_text", "")
        m = period_re.search(text)
        if m:
            start = m.group(1).rstrip(".")
            end   = m.group(2).rstrip(".")
            return f"{start} - {end}"
    return f"{EXTRACTION_DATE}"


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Fetch ssbbilla.site data
# ══════════════════════════════════════════════════════════════════════════════

def fetch_ssbbilla_products():
    """
    Download and parse the current ssbbilla.site brochure.
    Returns (list_of_products, extraction_date_str).
    """
    log.info("Fetching current ssbbilla.site data for comparison...")
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from billa_scraper import download_billa_page, parse_billa_html, parse_billa_markdown

        raw_path = download_billa_page(str(WORK_DIR / "ssbbilla_raw.html"))
        if not raw_path:
            log.warning("Could not download ssbbilla.site — comparison will be skipped")
            return [], EXTRACTION_DATE

        with open(raw_path, encoding="utf-8") as f:
            raw = f.read()

        is_html = '<div class="product">' in raw or '<span class="price">' in raw
        products = parse_billa_html(raw) if is_html else parse_billa_markdown(raw)
        log.info(f"ssbbilla.site: {len(products)} products parsed")
        return products, EXTRACTION_DATE

    except Exception as e:
        log.warning(f"Could not load ssbbilla.site data: {e}")
        return [], EXTRACTION_DATE


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — Compare PDF vs ssbbilla.site and generate report
# ══════════════════════════════════════════════════════════════════════════════

def compare_products(pdf_products, ssbbilla_products, threshold=COMPARISON_THRESHOLD):
    """
    Fuzzy-match each PDF product against all ssbbilla products.
    Returns list of row dicts for the comparison report.
    """
    rows = []
    for p in pdf_products:
        pdf_name  = p["product_name"]
        pdf_lower = pdf_name.lower()

        best_ratio  = 0.0
        best_match  = None
        for s in ssbbilla_products:
            ratio = SequenceMatcher(None, pdf_lower, s["product_name"].lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_match = s

        in_ss = best_ratio >= threshold
        rows.append({
            "pdf_product_name":       pdf_name,
            "ssbbilla_product_name":  best_match["product_name"] if best_match else "",
            "similarity_pct":         round(best_ratio * 100, 1),
            "pdf_promo_eur":          p["promo_price"],
            "ssbbilla_promo_eur":     best_match["promo_price"] if best_match else "",
            "price_diff_eur":         (
                round(p["promo_price"] - best_match["promo_price"], 2)
                if best_match and isinstance(best_match.get("promo_price"), float) else ""
            ),
            "in_ssbbilla":            "Yes" if in_ss else "No",
            "pdf_promo_period":       p.get("promo_period", ""),
            "ssbbilla_extraction_date": best_match.get("extraction_date", "") if best_match else "",
        })

    # Sort: unmatched items first, then by similarity ascending
    rows.sort(key=lambda r: (r["in_ssbbilla"] == "Yes", -r["similarity_pct"]))
    return rows


def save_comparison_report(rows, pdf_promo_period, ssbbilla_date, report_path):
    """Write the comparison report to an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed. Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # ── Summary row at top ────────────────────────────────────────────────────
    total     = len(rows)
    matched   = sum(1 for r in rows if r["in_ssbbilla"] == "Yes")
    unmatched = total - matched

    ws.append([f"Billa PDF vs ssbbilla.site — weekly comparison"])
    ws.append([f"PDF brochure period: {pdf_promo_period}"])
    ws.append([f"ssbbilla.site data pulled: {ssbbilla_date}"])
    ws.append([f"Report generated: {EXTRACTION_DATE}"])
    ws.append([
        f"Total PDF products: {total}  |  "
        f"Found on ssbbilla.site: {matched} ({100*matched//total if total else 0}%)  |  "
        f"PDF-only (not on ssbbilla): {unmatched}"
    ])
    ws.append([])  # blank separator

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "PDF Product Name",
        "ssbbilla.site Best Match",
        "Similarity %",
        "PDF Promo €",
        "ssbbilla Promo €",
        "Price Diff €",
        "In ssbbilla?",
        "PDF Promo Period",
        "ssbbilla Pull Date",
    ]
    header_row = ws.max_row + 1
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Data rows ─────────────────────────────────────────────────────────────
    green_fill  = PatternFill("solid", fgColor="C6EFCE")   # matched
    red_fill    = PatternFill("solid", fgColor="FFC7CE")    # not matched
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")    # matched but price differs

    for r in rows:
        price_diff = r["price_diff_eur"]
        row_data = [
            r["pdf_product_name"],
            r["ssbbilla_product_name"],
            r["similarity_pct"],
            r["pdf_promo_eur"],
            r["ssbbilla_promo_eur"],
            price_diff,
            r["in_ssbbilla"],
            r["pdf_promo_period"],
            r["ssbbilla_extraction_date"],
        ]
        ws.append(row_data)
        data_row = ws.max_row

        if r["in_ssbbilla"] == "Yes":
            fill = yellow_fill if (isinstance(price_diff, float) and abs(price_diff) > 0.01) else green_fill
        else:
            fill = red_fill

        for cell in ws[data_row]:
            cell.fill = fill

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [45, 45, 14, 12, 16, 12, 14, 22, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(str(report_path))
    log.info(f"Comparison report saved: {report_path}")
    log.info(f"  {matched}/{total} PDF products found on ssbbilla.site ({unmatched} PDF-only)")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 8 — Merge into master JSON
# ══════════════════════════════════════════════════════════════════════════════

def merge_into_master(new_items, source_label="Billa PDF-only"):
    """
    Add new_items into the master JSON.
    Existing Billa Direct records from the same promo_period are replaced.
    Items already present (by name + price dedup key) are skipped.
    """
    if not MASTER_PATH.exists():
        log.warning(f"Master JSON not found at {MASTER_PATH} — will create it")
        master = []
    else:
        with open(MASTER_PATH, encoding="utf-8") as f:
            master = json.load(f)

    before = len(master)
    # Remove Billa Direct records from this same promo period
    if new_items:
        period = new_items[0].get("promo_period", "")
        master = [
            r for r in master
            if not (
                r.get("source_store") == SOURCE_STORE
                and r.get("source_channel") == SOURCE_CHANNEL
                and r.get("promo_period") == period
            )
        ]
    removed = before - len(master)
    master.extend(new_items)

    seen = set()
    deduped = []
    for r in master:
        key = (
            r.get("source_store", "")[:15],
            r.get("source_channel", ""),
            r.get("product_name", "")[:40].lower(),
            r.get("promo_price"),
        )
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    with open(MASTER_PATH, "w", encoding="utf-8") as f:
        json.dump(deduped, f, ensure_ascii=False, indent=2)

    log.info(
        f"Master updated: removed {removed} old Billa Direct records for this period, "
        f"added {len(new_items)} {source_label} records, total {len(deduped)}"
    )
    return len(new_items)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def get_azure_key(args_key):
    """Resolve Azure key: CLI arg > secrets.py > env var."""
    if args_key:
        return args_key
    try:
        import secrets as _s
        if _s.AZURE_KEY and "your-azure" not in _s.AZURE_KEY:
            return _s.AZURE_KEY
    except ImportError:
        pass
    key = os.environ.get("AZURE_DI_KEY")
    if key:
        return key
    return None


def main():
    parser = argparse.ArgumentParser(
        description="Billa Bulgaria weekly PDF brochure pipeline",
        epilog="Example: python billa_pdf_pipeline.py --key YOUR_KEY",
    )
    parser.add_argument("--key",     help="Azure Document Intelligence API key")
    parser.add_argument("--pdf",     help="Path to already-downloaded PDF (skip download step)")
    parser.add_argument("--ocr-dir", help="Path to cached OCR output directory (skip OCR step)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and report only — do not write to master JSON")
    args = parser.parse_args()

    WORK_DIR.mkdir(parents=True, exist_ok=True)

    import requests
    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Determine PDF path ────────────────────────────────────────────────────
    pdf_path = None
    promo_period = EXTRACTION_DATE  # default until OCR tells us otherwise

    if args.ocr_dir:
        log.info(f"Reusing cached OCR from: {args.ocr_dir}")
    else:
        if args.pdf:
            pdf_path = Path(args.pdf)
            if not pdf_path.exists():
                log.error(f"PDF not found: {pdf_path}")
                sys.exit(1)
            log.info(f"Using provided PDF: {pdf_path}")
        else:
            # Auto-download
            viewer_url, slug = find_publitas_viewer_url(session)
            if not viewer_url:
                log.error("Could not determine Publitas viewer URL. "
                          "Use --pdf to provide the PDF manually.")
                sys.exit(1)

            pdf_filename = f"billa_brochure_{EXTRACTION_DATE}.pdf"
            pdf_path = WORK_DIR / pdf_filename

            if pdf_path.exists():
                log.info(f"PDF already exists locally: {pdf_path} — skipping download")
            else:
                pdf_path_result = download_pdf(viewer_url, str(pdf_path), session)
                if not pdf_path_result:
                    log.error(
                        "PDF download failed.\n"
                        "Manual fallback:\n"
                        f"  1. Open {viewer_url}\n"
                        f"  2. Click 'Изтегляне на PDF'\n"
                        f"  3. Save to: {pdf_path}\n"
                        f"  4. Re-run: python billa_pdf_pipeline.py --pdf {pdf_path}"
                    )
                    sys.exit(1)

        # ── Split & OCR ───────────────────────────────────────────────────────
        azure_key = get_azure_key(args.key)
        if not azure_key:
            log.error(
                "No Azure key provided.\n"
                "Options:\n"
                "  1. Pass --key YOUR_KEY\n"
                "  2. Set AZURE_KEY in secrets.py\n"
                "  3. Set env var AZURE_DI_KEY"
            )
            sys.exit(1)

        batch_dir = WORK_DIR / "pdf_batches"
        ocr_dir   = WORK_DIR / "ocr_output"

        batches = split_pdf(pdf_path, batch_dir)
        ocr_results = ocr_all_batches(batches, ocr_dir, AZURE_ENDPOINT, azure_key)

    if args.ocr_dir:
        ocr_results = load_ocr_from_directory(args.ocr_dir)
        if not ocr_results:
            sys.exit(1)

    # ── Parse OCR ─────────────────────────────────────────────────────────────
    promo_period = extract_promo_period(ocr_results)
    log.info(f"Brochure promo period: {promo_period}")

    pdf_products = parse_ocr_results(ocr_results, promo_period)
    log.info(f"Parsed {len(pdf_products)} products from PDF OCR")

    if not pdf_products:
        log.error("No products extracted from OCR output — check the OCR cache.")
        sys.exit(1)

    # ── Comparison (if enabled) ───────────────────────────────────────────────
    items_to_merge = pdf_products  # default: merge everything from PDF

    if WEEKLY_COMPARISON:
        ssbbilla_products, ss_date = fetch_ssbbilla_products()

        if ssbbilla_products:
            rows = compare_products(pdf_products, ssbbilla_products)

            report_path = WORK_DIR / f"comparison_{EXTRACTION_DATE}.xlsx"
            save_comparison_report(rows, promo_period, ss_date, report_path)

            # Only supplement master with items NOT found on ssbbilla.site
            items_to_merge = [
                pdf_products[i] for i, r in enumerate(rows)
                if r["in_ssbbilla"] == "No"
            ]
            log.info(
                f"Comparison done: {len(items_to_merge)} PDF-only items "
                f"will be added to master (not on ssbbilla.site)"
            )
        else:
            log.warning("ssbbilla.site data unavailable — merging all PDF items")
    else:
        log.info("BILLA_WEEKLY_COMPARISON disabled — merging all PDF items into master")

    # ── Merge into master ─────────────────────────────────────────────────────
    if args.dry_run:
        log.info(f"DRY RUN — not writing to master. Would add {len(items_to_merge)} items.")
        for p in items_to_merge[:10]:
            log.info(f"  [{p['product_category']}] {p['product_name']} | "
                     f"promo={p['promo_price']} € reg={p['regular_price']} €")
    else:
        label = "PDF-only" if WEEKLY_COMPARISON else "PDF"
        merge_into_master(items_to_merge, source_label=label)


if __name__ == "__main__":
    main()
